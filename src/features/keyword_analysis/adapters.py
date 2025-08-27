"""
벤더 정규화 → 기능형 데이터로 변환 + 엑셀 저장
네이버 API 응답을 키워드 분석 전용 데이터로 가공하고 엑셀로 내보내기
"""
from typing import List, Dict, Any, Optional
from collections import Counter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from src.foundation.logging import get_logger
from src.foundation.exceptions import FileError
from src.foundation.http_client import default_http_client, api_error_handler
from src.features.keyword_analysis.models import KeywordData, AnalysisResult
from src.features.keyword_analysis.engine_local import calculate_competition_strength

logger = get_logger("features.keyword_analysis.adapters")


class KeywordAnalysisAdapter:
    """키워드 분석 데이터 어댑터"""
    
    @staticmethod
    def extract_search_volume(searchad_data: Dict[str, Any], keyword: str) -> Optional[int]:
        """
        검색광고 데이터에서 검색량 추출 (raw/정규화 응답 모두 처리)
        
        Args:
            searchad_data: 검색광고 API 응답 (raw 또는 정규화됨)
            keyword: 대상 키워드
        
        Returns:
            Optional[int]: 월간 검색량
        """
        try:
            # 정규화된 응답 처리
            keywords = searchad_data.get('keywords', [])
            if keywords:
                # 정확히 일치하는 키워드 찾기
                for kw_data in keywords:
                    if kw_data.get('keyword', '').strip().upper() == keyword.strip().upper():
                        return kw_data.get('monthly_total_searches', 0)
                
                # 일치하는 키워드가 없으면 첫 번째 키워드의 검색량 사용
                if keywords:
                    return keywords[0].get('monthly_total_searches', 0)
            
            # Raw 응답 처리 (keywordList 필드)
            keyword_list = searchad_data.get('keywordList', [])
            if keyword_list:
                # 첫 번째 키워드의 검색량 추출
                first_keyword = keyword_list[0]
                
                # "< 10" 값을 0으로 처리
                pc_count = first_keyword.get('monthlyPcQcCnt', '0')
                mobile_count = first_keyword.get('monthlyMobileQcCnt', '0')
                
                monthly_pc = 0 if pc_count == '< 10' else int(pc_count or 0)
                monthly_mobile = 0 if mobile_count == '< 10' else int(mobile_count or 0)
                
                return monthly_pc + monthly_mobile
            
            return None
            
        except Exception as e:
            logger.warning(f"검색량 추출 실패 - {keyword}: {e}")
            return None
    
    @staticmethod
    def extract_category_for_keyword_analysis(shopping_data: Dict[str, Any]) -> str:
        """
        쇼핑 데이터에서 키워드 분석용 카테고리 추출
        상위 40개 상품의 카테고리를 분석하여 전체 경로로 반환
        
        Args:
            shopping_data: 정규화된 쇼핑 API 응답
        
        Returns:
            str: 대표 카테고리 전체 경로 (예: "패션의류 > 여성의류 > 원피스")
        """
        try:
            products = shopping_data.get('products', [])
            if not products:
                return ""
            
            # 모든 상품의 카테고리 경로 수집
            all_category_paths = []
            for product in products:
                categories = product.get('categories', [])
                if categories:
                    # 카테고리를 '/'로 연결하여 전체 경로 생성 (기존 통합관리프로그램과 동일)
                    category_path = '/'.join(categories)
                    all_category_paths.append(category_path)
            
            if not all_category_paths:
                return ""
            
            # 상위 2개 카테고리 경로 찾기 (기존 통합관리프로그램과 동일)
            category_counter = Counter(all_category_paths)
            most_common = category_counter.most_common(2)  # 상위 2개
            
            if most_common:
                total = len(all_category_paths)
                result_lines = []
                
                for category_path, count in most_common:
                    percentage = int((count / total) * 100)
                    result_lines.append(f"{category_path}({percentage}%)")
                
                result = "\n".join(result_lines)
                logger.debug(f"카테고리 분석 결과: {len(most_common)}개 카테고리")
                return result
            
            return ""
            
        except Exception as e:
            logger.warning(f"카테고리 추출 실패: {e}")
            return ""
    
    @staticmethod
    def extract_total_products(shopping_data: Dict[str, Any]) -> Optional[int]:
        """
        쇼핑 데이터에서 총 상품 수 추출
        
        Args:
            shopping_data: 정규화된 쇼핑 API 응답
        
        Returns:
            Optional[int]: 총 상품 수
        """
        try:
            return shopping_data.get('total_count', 0)
        except Exception as e:
            logger.warning(f"상품 수 추출 실패: {e}")
            return None
    
    
    @staticmethod
    def build_keyword_data(keyword: str,
                          searchad_data: Optional[Dict[str, Any]] = None,
                          shopping_data: Optional[Dict[str, Any]] = None) -> KeywordData:
        """
        키워드 분석 데이터 구성
        
        Args:
            keyword: 키워드
            searchad_data: 정규화된 검색광고 API 응답
            shopping_data: 정규화된 쇼핑 API 응답
        
        Returns:
            KeywordData: 키워드 분석 데이터
        """
        try:
            # 검색량 추출
            search_volume = None
            if searchad_data:
                search_volume = KeywordAnalysisAdapter.extract_search_volume(searchad_data, keyword)
            
            # 카테고리 및 상품 수 추출
            category = ""
            total_products = None
            if shopping_data:
                category = KeywordAnalysisAdapter.extract_category_for_keyword_analysis(shopping_data)
                total_products = KeywordAnalysisAdapter.extract_total_products(shopping_data)
            
            # 경쟁 강도 계산 (engine_local 사용)
            competition_strength = calculate_competition_strength(
                search_volume, total_products
            )
            
            keyword_data = KeywordData(
                keyword=keyword,
                category=category,
                search_volume=search_volume,
                total_products=total_products,
                competition_strength=competition_strength
            )
            
            logger.debug(f"키워드 데이터 구성 완료: {keyword}")
            return keyword_data
            
        except Exception as e:
            logger.error(f"키워드 데이터 구성 실패 - {keyword}: {e}")
            # 오류 발생 시 기본 데이터 반환
            return KeywordData(keyword=keyword)



# 편의 함수들
def adapt_keyword_data(keyword: str,
                      searchad_data: Optional[Dict[str, Any]] = None,
                      shopping_data: Optional[Dict[str, Any]] = None) -> KeywordData:
    """키워드 데이터 어댑터 편의 함수"""
    return KeywordAnalysisAdapter.build_keyword_data(keyword, searchad_data, shopping_data)




class KeywordExcelAdapter:
    """키워드 분석 엑셀 내보내기 어댑터"""
    
    def __init__(self):
        self.default_font = Font(name='맑은 고딕', size=10)
        self.header_font = Font(name='맑은 고딕', size=11, bold=True)
        self.header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def export_analysis_result(self, result: AnalysisResult, file_path: str) -> bool:
        """
        키워드 분석 결과를 엑셀로 내보내기
        
        Args:
            result: 키워드 분석 결과
            file_path: 저장할 파일 경로
        
        Returns:
            bool: 성공 여부
        """
        try:
            if not result.keywords:
                logger.warning("내보낼 키워드 데이터가 없습니다")
                return False
            
            # KeywordData를 딕셔너리로 변환
            data = []
            for kw in result.keywords:
                data.append({
                    'keyword': kw.keyword,
                    'category': kw.category,
                    'search_volume': kw.search_volume,
                    'total_products': kw.total_products,
                    'competition_strength': kw.competition_strength
                })
            
            return self.export_keywords(data, file_path)
            
        except Exception as e:
            logger.error(f"분석 결과 엑셀 내보내기 실패: {e}")
            raise FileError(f"엑셀 내보내기 실패: {e}")
    
    def export_keywords(self, 
                       data: List[Dict[str, Any]], 
                       file_path: str,
                       sheet_name: str = "키워드 분석") -> bool:
        """
        키워드 데이터를 Excel로 내보내기
        
        Args:
            data: 키워드 데이터 리스트
            file_path: 저장할 파일 경로
            sheet_name: 시트 이름
        
        Returns:
            bool: 성공 여부
        """
        try:
            if not data:
                logger.warning("내보낼 데이터가 없습니다")
                return False
            
            # DataFrame 생성 전 Excel 안전화 (NaN/Inf 처리)
            safe_data = []
            for row in data:
                safe_row = {}
                for key, value in row.items():
                    if isinstance(value, float):
                        import math
                        if math.isnan(value) or math.isinf(value):
                            safe_row[key] = None  # Excel에서 빈 셀로 표시
                        else:
                            safe_row[key] = value
                    else:
                        safe_row[key] = value
                safe_data.append(safe_row)
            
            # DataFrame 생성
            df = pd.DataFrame(safe_data)
            
            # 컬럼명 한글화
            column_mapping = {
                'keyword': '키워드',
                'category': '카테고리',
                'search_volume': '월간 검색량',
                'total_products': '상품 수',
                'competition_strength': '경쟁 강도'
            }
            
            df = df.rename(columns=column_mapping)
            
            # Excel 파일 생성
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 데이터 추가
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # 스타일 적용
            self._apply_keyword_styles(ws, len(df.columns))
            
            # 파일 저장
            wb.save(file_path)
            logger.info(f"키워드 분석 엑셀 파일 저장 완료: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"키워드 엑셀 내보내기 실패: {e}")
            raise FileError(f"엑셀 파일 생성 실패: {e}")
    
    def _apply_keyword_styles(self, worksheet, column_count: int):
        """키워드 분석용 워크시트 스타일 적용"""
        try:
            # 헤더 스타일
            for col in range(1, column_count + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
            
            # 데이터 스타일
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = self.default_font
                    col_idx = cell.col_idx  # 문자(A,B) 대신 숫자 인덱스 사용
                    if col_idx == 1:  # 키워드 - 좌측 정렬
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif col_idx == 2:  # 카테고리 - 줄바꿈 허용
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    elif col_idx == 3:  # 월간 검색량 - 천단위 콤마
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'
                    elif col_idx == 4:  # 상품 수 - 천단위 콤마
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'
                    elif col_idx == 5:  # 경쟁 강도 - 소수점 2자리
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00'
                    else:
                        cell.alignment = self.center_alignment
            
            # 컬럼 너비 설정
            column_widths = {
                1: 20,   # 키워드
                2: 50,   # 카테고리 (줄바꿈을 위해 넓게)
                3: 15,   # 월간 검색량
                4: 15,   # 상품 수
                5: 12    # 경쟁 강도
            }
            
            for col_num, width in column_widths.items():
                if col_num <= column_count:
                    column_letter = worksheet.cell(row=1, column=col_num).column_letter
                    worksheet.column_dimensions[column_letter].width = width
            
            # 행 높이 설정 (카테고리 줄바꿈을 위해)
            for row_num in range(2, worksheet.max_row + 1):
                category_cell = worksheet.cell(row=row_num, column=2)
                if category_cell.value and '\n' in str(category_cell.value):
                    worksheet.row_dimensions[row_num].height = 30
                
        except Exception as e:
            logger.warning(f"키워드 분석 스타일 적용 중 오류: {e}")


# 편의 함수들
def export_keywords_to_excel(keywords: List[KeywordData], file_path: str) -> bool:
    """키워드 리스트를 엑셀로 내보내기"""
    try:
        data = []
        for kw in keywords:
            data.append({
                'keyword': kw.keyword,
                'category': kw.category or '-',
                'search_volume': kw.search_volume,
                'total_products': kw.total_products,
                'competition_strength': kw.competition_strength
            })
        
        adapter = KeywordExcelAdapter()
        return adapter.export_keywords(data, file_path)
        
    except Exception as e:
        logger.error(f"키워드 엑셀 내보내기 실패: {e}")
        return False


# TODO[adapters]: service.py에서 요청한 함수들 추가 필요
@api_error_handler("네이버 검색광고 API")
def fetch_searchad_raw(keyword: str) -> Optional[Dict[str, Any]]:
    """검색광고 API Raw 데이터 수집 - Foundation HTTP Client 사용"""
    try:
        from src.vendors.naver.client_factory import get_keyword_client
        client = get_keyword_client()
        if client:
            # vendors/naver/searchad/keyword_client.py 표준 (내부적으로 foundation HTTP 사용하도록 수정 예정)
            return client.get_keyword_ideas([keyword])
        return None
    except Exception as e:
        logger.warning(f"검색광고 데이터 수집 실패 - {keyword}: {e}")
        return None

@api_error_handler("네이버 쇼핑 API")
def fetch_shopping_normalized(keyword: str) -> Optional[Dict[str, Any]]:
    """쇼핑 API 정규화 데이터 수집 - Foundation HTTP Client 사용"""
    try:
        from src.vendors.naver.client_factory import get_shopping_client
        client = get_shopping_client()
        if client:
            # vendors/naver/developer/shopping_client.py 표준 (내부적으로 foundation HTTP 사용하도록 수정 예정)
            raw = client.search_products(query=keyword, display=40, sort="sim")
            from src.vendors.naver.normalizers import normalize_shopping_response
            return normalize_shopping_response(raw)
        return None
    except Exception as e:
        logger.warning(f"쇼핑 데이터 수집 실패 - {keyword}: {e}")
        return None


def export_analysis_result_to_excel(result: AnalysisResult, file_path: str) -> bool:
    """분석 결과를 엑셀로 내보내기"""
    try:
        adapter = KeywordExcelAdapter()
        return adapter.export_analysis_result(result, file_path)
        
    except Exception as e:
        logger.error(f"분석 결과 엑셀 내보내기 실패: {e}")
        return False