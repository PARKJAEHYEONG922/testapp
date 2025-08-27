"""
ìˆœìœ„ ì¶”ì  ì–´ëŒ‘í„° - vendors ë ˆì´ì–´ ì‘ë‹µì„ features ë°ì´í„°ë¡œ ê°€ê³µ
Raw API ì‘ë‹µì„ ë¹„ì¦ˆë‹ˆìŠ¤ ë¡œì§ì—ì„œ ì‚¬ìš©í•  ìˆ˜ ìˆëŠ” í˜•íƒœë¡œ ë³€í™˜
ì—‘ì…€ ë‚´ë³´ë‚´ê¸° í¬í•¨
"""
from typing import Optional, Dict, Any, List, TypedDict, Tuple
import re
from datetime import datetime

from src.vendors.naver.developer.shopping_client import shopping_client as naver_shopping
from src.vendors.naver.client_factory import get_keyword_tool_client
from src.toolbox.text_utils import validate_naver_url, extract_product_id, validate_product_id
from src.toolbox.formatters import format_monthly_volume, format_rank, format_datetime, format_price_krw
from src.foundation.logging import get_logger
from .models import RANK_OUT_OF_RANGE

logger = get_logger("features.rank_tracking.adapters")


class ProductInfoDTO(TypedDict, total=False):
    """ìƒí’ˆ ì •ë³´ DTO"""
    product_id: str
    name: str
    price: int
    category: str
    store_name: str
    description: str
    image_url: str
    url: str


class RankingCheckDTO(TypedDict, total=False):
    """ìˆœìœ„ í™•ì¸ ê²°ê³¼ DTO"""
    success: bool
    rank: int
    total_results: int
    error: str
    keyword: str
    product_id: str


def format_date(date_str: str) -> str:
    """ë‚ ì§œ í˜•ì‹ ë³€í™˜ (8/6 14:26)"""
    dt = _to_dt(date_str)
    return dt.strftime("%m/%d %H:%M") if dt else date_str


def format_date_with_time(date_str: str) -> str:
    """ë‚ ì§œ ì‹œê°„ í˜•ì‹ ë³€í™˜ (2025-08-07 15:23)"""
    dt = _to_dt(date_str)
    return dt.strftime("%Y-%m-%d %H:%M") if dt else date_str


def format_rank_display(rank: int) -> str:
    """ìˆœìœ„ ìˆ«ìë¥¼ ì‚¬ìš©ì ì¹œí™”ì ì¸ í˜•íƒœë¡œ í¬ë§·íŒ… (UI í‘œì‹œìš©)"""
    if not isinstance(rank, int):
        return "-"
    if rank == RANK_OUT_OF_RANGE or rank > 200:
        return "200ìœ„ë°–"
    elif rank >= 1:
        return f"{rank}ìœ„"
    return "-"


def _to_dt(date_str: str):
    """ë¬¸ìì—´ì„ datetime ê°ì²´ë¡œ ë³€í™˜ (DATEì™€ ISO í˜•ì‹ ëª¨ë‘ ì²˜ë¦¬)"""
    try:
        # 'YYYY-MM-DD HH:MM:SS' ë˜ëŠ” 'YYYY-MM-DDTHH:MM:SS[Z]' í˜•ì‹
        return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
    except:
        try:
            # 'YYYY-MM-DD HH:MM:SS' í˜•ì‹ (workerì—ì„œ ìƒì„±)
            return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                # 'YYYY-MM-DD' í˜•ì‹ (repositoryì—ì„œ ë°˜í™˜)
                return datetime.strptime(date_str, "%Y-%m-%d")
            except:
                return None




def get_rank_color(rank: int, color_type: str = "background") -> str:
    """ìˆœìœ„ì— ë”°ë¥¸ ìƒ‰ìƒ ë°˜í™˜"""
    from src.foundation.logging import get_logger
    logger = get_logger("rank_color_debug")
    
    # ë””ë²„ê¹…: 200ìœ„ë°– ì¼€ì´ìŠ¤ ë¡œê¹…
    if rank > 200 or rank == 999:
        logger.info(f"ğŸ”´ 200ìœ„ë°– ê°ì§€! rank={rank}, type={type(rank)}, color_type={color_type}")
    
    if color_type == "background":
        # ë°°ê²½ìƒ‰ (ì—°í•œ í†¤)
        if rank <= 10:
            return "#e8f5e8"  # ì—°í•œ ì´ˆë¡
        elif rank <= 40:
            return "#fff3cd"  # ì—°í•œ ë…¸ë‘
        else:
            return "#f8d7da"  # ì—°í•œ ë¹¨ê°•
    else:  # foreground/text color
        # í…ìŠ¤íŠ¸ ìƒ‰ìƒ (ì§„í•œ í†¤)
        if rank == -1 or rank == 0:  # ê²€ìƒ‰ëŸ‰ ì—†ìŒ/API ì‹¤íŒ¨
            logger.info(f"ğŸ”˜ API ì‹¤íŒ¨/ê²€ìƒ‰ëŸ‰ ì—†ìŒ: rank={rank} -> íšŒìƒ‰")
            return "#6B7280"  # íšŒìƒ‰
        elif rank <= 10:
            logger.info(f"ğŸŸ¢ 1-10ìœ„: rank={rank} -> ì´ˆë¡ìƒ‰")
            return "#059669"  # ì´ˆë¡ìƒ‰ (ìƒìœ„ 10ìœ„)
        elif rank <= 40:
            logger.info(f"ğŸŸ¡ 11-40ìœ„: rank={rank} -> ì£¼í™©ìƒ‰")
            return "#D97706"  # ì£¼í™©ìƒ‰ (40ìœ„ ì´ë‚´)
        elif rank > 200 or rank == 999:  # 200ìœ„ë°– ëª…ì‹œì  ì²˜ë¦¬
            logger.info(f"ğŸ”´ 200ìœ„ë°–: rank={rank} -> ë” ì§„í•œ ë¹¨ê°„ìƒ‰ #B91C1C")
            return "#B91C1C"  # ë” ì§„í•œ ë¹¨ê°„ìƒ‰ (200ìœ„ë°–)
        else:
            logger.info(f"ğŸ”´ 41-200ìœ„: rank={rank} -> ë¹¨ê°„ìƒ‰")
            return "#DC2626"  # ë¹¨ê°„ìƒ‰ (41-200ìœ„)


# ê¸°ì¡´ format_monthly_volumeì€ ì‚­ì œë¨ - toolbox.formatters.format_monthly_volume ì‚¬ìš©


def get_category_match_color(project_category: str, keyword_category: str) -> str:
    """ì¹´í…Œê³ ë¦¬ ë§¤ì¹­ ê²°ê³¼ì— ë”°ë¥¸ ìƒ‰ìƒ ë°˜í™˜"""
    if not project_category or not keyword_category:
        return "#6B7280"  # íšŒìƒ‰ (ë°ì´í„° ì—†ìŒ)
    
    # ì „ì²´ ì¹´í…Œê³ ë¦¬ ê²½ë¡œ ë¹„êµ - ê´„í˜¸ ë¶€ë¶„ ì œê±° í›„ ì „ì²´ ë¹„êµ
    project_clean = project_category.split('(')[0].strip() if '(' in project_category else project_category
    keyword_clean = keyword_category.split('(')[0].strip() if '(' in keyword_category else keyword_category
    
    # ì™„ì „ ì¼ì¹˜ë§Œ ì´ˆë¡ìƒ‰ (ì „ì²´ ê²½ë¡œê°€ ì™„ì „íˆ ë™ì¼í•´ì•¼ í•¨)
    if project_clean == keyword_clean:
        return "#059669"  # ì´ˆë¡ìƒ‰ (ì™„ì „ ì¼ì¹˜)
    else:
        return "#DC2626"  # ë¹¨ê°„ìƒ‰ (ë¶ˆì¼ì¹˜ - ë¶€ë¶„ì¼ì¹˜ í¬í•¨)


def clean_product_name(name: str) -> str:
    """ìƒí’ˆëª… ì •ë¦¬ (ê³µë°± ì •ê·œí™”) - serviceì—ì„œ ì‚¬ìš© ê°€ëŠ¥"""
    if not name:
        return ""
    
    import re
    # ì—°ì†ëœ ê³µë°±ì„ í•˜ë‚˜ë¡œ í†µì¼ ë° ì•ë’¤ ê³µë°± ì œê±°
    clean_name = re.sub(r'\s+', ' ', name).strip()
    
    return clean_name


def smart_product_search(product_name: str, product_id: str) -> Optional[Dict[str, Any]]:
    """ìƒí’ˆ ìŠ¤ë§ˆíŠ¸ ê²€ìƒ‰ - serviceì—ì„œ ì‚¬ìš© ê°€ëŠ¥"""
    try:
        return naver_shopping.smart_product_search(product_name, product_id)
    except Exception as e:
        logger.error(f"ìŠ¤ë§ˆíŠ¸ ìƒí’ˆ ê²€ìƒ‰ ì‹¤íŒ¨: {product_name} ({product_id}): {e}")
        return None


class RankTrackingAdapter:
    """ìˆœìœ„ ì¶”ì  ì–´ëŒ‘í„°"""
    
    def __init__(self):
        self.shopping_client = naver_shopping
        self.keyword_client = get_keyword_tool_client()
    
    def extract_product_id_from_url(self, url: str) -> str:
        """ë„¤ì´ë²„ ì‡¼í•‘ URLì—ì„œ ìƒí’ˆ ID ì¶”ì¶œ (validators ì‚¬ìš©)"""
        if not url or not isinstance(url, str):
            raise ValueError("URLì´ ë¹„ì–´ìˆê±°ë‚˜ ì˜¬ë°”ë¥´ì§€ ì•ŠìŠµë‹ˆë‹¤")
        
        if not validate_naver_url(url):
            raise ValueError(
                f"ì§€ì›ë˜ì§€ ì•ŠëŠ” ë„¤ì´ë²„ ì‡¼í•‘ URL í˜•ì‹ì…ë‹ˆë‹¤: {url}\n"
                "ì˜¬ë°”ë¥¸ í˜•ì‹: https://shopping.naver.com/catalog/XXXXX ë˜ëŠ” "
                "https://smartstore.naver.com/store/products/XXXXX"
            )
        
        product_id = extract_product_id(url)
        if not product_id:
            raise ValueError(f"URLì—ì„œ ìƒí’ˆ IDë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {url}")
        
        if not validate_product_id(product_id):
            raise ValueError(f"ìœ íš¨í•˜ì§€ ì•Šì€ ìƒí’ˆ IDì…ë‹ˆë‹¤: {product_id}")
        
        return product_id
    
    def get_product_info(self, product_name: str, product_id: str) -> Optional[ProductInfoDTO]:
        """ìƒí’ˆ ì •ë³´ ì¡°íšŒ (vendors -> ProductInfoDTO ë³€í™˜)"""
        try:
            raw_data = self.shopping_client.smart_product_search(product_name, product_id)
            if not raw_data:
                return None
            
            return ProductInfoDTO(
                product_id=raw_data.get('product_id', ''),
                name=self._clean_product_name(raw_data.get('name', '')),
                price=raw_data.get('price', 0),
                category=raw_data.get('category', ''),  # ì „ì²´ ì¹´í…Œê³ ë¦¬ ê²½ë¡œ ìœ ì§€
                store_name=raw_data.get('store_name', ''),
                description=raw_data.get('description', ''),
                image_url=raw_data.get('image_url', ''),
                url=raw_data.get('url', '')
            )
            
        except Exception as e:
            logger.error(f"ìƒí’ˆ ì •ë³´ ì¡°íšŒ ì‹¤íŒ¨: {product_name} ({product_id}): {e}")
            return None
    
    def check_product_rank(self, keyword: str, product_id: str) -> RankingCheckDTO:
        """í‚¤ì›Œë“œì—ì„œ ìƒí’ˆ ìˆœìœ„ í™•ì¸"""
        try:
            rank = self.shopping_client.find_product_rank(keyword, product_id, max_pages=2)
            
            result = RankingCheckDTO(
                success=True,
                rank=rank if rank is not None else RANK_OUT_OF_RANGE,
                total_results=max(rank, 100) if rank is not None else 1000,
                keyword=keyword,
                product_id=product_id
            )
            
            logger.info(f"ìˆœìœ„ í™•ì¸ ì„±ê³µ: {keyword} -> {product_id} = {rank or '200+'}ìœ„")
            return result
            
        except Exception as e:
            logger.error(f"ìˆœìœ„ í™•ì¸ ì‹¤íŒ¨: {keyword} -> {product_id}: {e}")
            return RankingCheckDTO(
                success=False,
                rank=RANK_OUT_OF_RANGE,
                total_results=0,
                error=str(e),
                keyword=keyword,
                product_id=product_id
            )
    
    def get_keyword_monthly_volume(self, keyword: str) -> Optional[int]:
        """í‚¤ì›Œë“œ ì›” ê²€ìƒ‰ëŸ‰ ì¡°íšŒ (ê²€ìƒ‰ê´‘ê³  API í™œìš©)"""
        try:
            volume = self.keyword_client.get_single_search_volume(keyword)
            logger.debug(f"ì›”ê²€ìƒ‰ëŸ‰ ì¡°íšŒ: {keyword} -> {volume}")
            return volume
        except Exception as e:
            logger.warning(f"ì›”ê²€ìƒ‰ëŸ‰ ì¡°íšŒ ì‹¤íŒ¨: {keyword}: {e}")
            return None
    
    def get_keyword_category(self, keyword: str, sample_size: int = None) -> Optional[str]:
        """í‚¤ì›Œë“œ ëŒ€í‘œ ì¹´í…Œê³ ë¦¬ ì¡°íšŒ (ì‡¼í•‘ API í™œìš©)"""
        try:
            if sample_size is None:
                from .models import DEFAULT_SAMPLE_SIZE
                sample_size = DEFAULT_SAMPLE_SIZE
            category = self.shopping_client.get_keyword_category(keyword, sample_size=sample_size)
            logger.debug(f"ì¹´í…Œê³ ë¦¬ ì¡°íšŒ: {keyword} -> {category}")
            return category
        except Exception as e:
            logger.warning(f"ì¹´í…Œê³ ë¦¬ ì¡°íšŒ ì‹¤íŒ¨: {keyword}: {e}")
            return None
    
    def analyze_keyword_for_tracking(self, keyword: str) -> Dict[str, Any]:
        """ì¶”ì ìš© í‚¤ì›Œë“œ ì¢…í•© ë¶„ì„ (ì›”ê²€ìƒ‰ëŸ‰ + ì¹´í…Œê³ ë¦¬)"""
        result = {
            'keyword': keyword,
            'monthly_volume': -1,
            'category': '',
            'success': False,
            'error_message': None
        }
        
        try:
            # ì›”ê²€ìƒ‰ëŸ‰ ì¡°íšŒ
            monthly_volume = self.get_keyword_monthly_volume(keyword)
            if monthly_volume is not None:
                result['monthly_volume'] = monthly_volume
            
            # ì¹´í…Œê³ ë¦¬ ì¡°íšŒ
            category = self.get_keyword_category(keyword)
            if category:
                result['category'] = category
            
            result['success'] = True
            logger.info(f"í‚¤ì›Œë“œ ë¶„ì„ ì™„ë£Œ: {keyword} (ë³¼ë¥¨: {monthly_volume}, ì¹´í…Œê³ ë¦¬: {category})")
            
        except Exception as e:
            result['error_message'] = str(e)
            logger.error(f"í‚¤ì›Œë“œ ë¶„ì„ ì‹¤íŒ¨: {keyword}: {e}")
        
        return result
    
    
    def check_multiple_keywords_rank(self, keywords: List[str], product_id: str) -> List[RankingCheckDTO]:
        """ì—¬ëŸ¬ í‚¤ì›Œë“œì˜ ìˆœìœ„ë¥¼ í•œë²ˆì— ê²€ìƒ‰"""
        results = []
        for keyword in keywords:
            try:
                result = self.check_product_rank(keyword, product_id)
                results.append(result)
            except Exception as e:
                logger.error(f"í‚¤ì›Œë“œ ìˆœìœ„ ê²€ìƒ‰ ì‹¤íŒ¨: {keyword}: {e}")
                # ì‹¤íŒ¨í•œ ê²½ìš°ë„ ê²°ê³¼ì— í¬í•¨
                failed_result = RankingCheckDTO(
                    success=False,
                    rank=RANK_OUT_OF_RANGE,
                    total_results=0,
                    error=str(e),
                    keyword=keyword,
                    product_id=product_id
                )
                results.append(failed_result)
        
        return results
    
    def analyze_keywords_for_tracking(self, keywords: List[str]) -> List[Dict[str, Any]]:
        """ì—¬ëŸ¬ í‚¤ì›Œë“œì˜ ê²€ìƒ‰ëŸ‰ê³¼ ì¹´í…Œê³ ë¦¬ë¥¼ í•œë²ˆì— ë¶„ì„"""
        results = []
        for keyword in keywords:
            try:
                result = self.analyze_keyword_for_tracking(keyword)
                results.append(result)
            except Exception as e:
                logger.error(f"í‚¤ì›Œë“œ ë¶„ì„ ì‹¤íŒ¨: {keyword}: {e}")
                # ì‹¤íŒ¨í•œ ê²½ìš°ë„ ê²°ê³¼ì— í¬í•¨
                failed_result = {
                    'keyword': keyword,
                    'monthly_volume': -1,
                    'category': '',
                    'success': False,
                    'error_message': str(e)
                }
                results.append(failed_result)
        
        return results
    
    def _clean_product_name(self, name: str) -> str:
        """ìƒí’ˆëª… ì •ë¦¬ (HTML íƒœê·¸ ì œê±° ë“±)"""
        return clean_product_name(name)
    
    def check_keyword_ranking(self, keyword: str, product_id: str) -> RankingCheckDTO:
        """í‚¤ì›Œë“œ ìˆœìœ„ í™•ì¸ (í˜¸í˜¸ì„± ìœ„í•œ alias - í–¥í›„ check_product_rankë¡œ ë§ˆì´ê·¸ë ˆì´ì…˜ í›„ ì œê±°)"""
        return self.check_product_rank(keyword, product_id)
    
    def analyze_keywords_batch(self, keywords: List[str]) -> Dict[str, Any]:
        """í‚¤ì›Œë“œ ë°°ì¹˜ ì›”ê²€ìƒ‰ëŸ‰ ë¶„ì„ (engine_localì—ì„œ ì´ë™)"""
        try:
            updated_count = 0
            failed_count = 0
            results = []
            
            for keyword in keywords:
                try:
                    analysis = self.analyze_keyword_for_tracking(keyword)
                    
                    if analysis['success']:
                        updated_count += 1
                        results.append({
                            'keyword': keyword,
                            'success': True,
                            'category': analysis['category'],
                            'monthly_volume': analysis['monthly_volume']
                        })
                    else:
                        failed_count += 1
                        results.append({
                            'keyword': keyword,
                            'success': False,
                            'error_message': analysis.get('error_message', 'ë¶„ì„ ì‹¤íŒ¨')
                        })
                        
                except Exception as e:
                    failed_count += 1
                    results.append({
                        'keyword': keyword,
                        'success': False,
                        'error_message': str(e)
                    })
                    logger.error(f"í‚¤ì›Œë“œ '{keyword}' ì²˜ë¦¬ ì‹¤íŒ¨: {e}")
            
            return {
                'success': updated_count > 0,
                'updated_count': updated_count,
                'failed_count': failed_count,
                'total_count': len(keywords),
                'results': results
            }
            
        except Exception as e:
            logger.error(f"í‚¤ì›Œë“œ ë°°ì¹˜ ë¶„ì„ ì‹¤íŒ¨: {e}")
            return {
                'success': False,
                'updated_count': 0,
                'failed_count': len(keywords),
                'total_count': len(keywords),
                'results': [],
                'error_message': str(e)
            }
    
    def check_project_rankings_analysis(self, project, keywords: List) -> Dict[str, Any]:
        """í”„ë¡œì íŠ¸ ì „ì²´ í‚¤ì›Œë“œ ìˆœìœ„ í™•ì¸ ë¶„ì„ (engine_localì—ì„œ ì´ë™)"""
        try:
            # ìˆœìœ„ í™•ì¸ ê²°ê³¼
            results = []
            success_count = 0
            
            for keyword_obj in keywords:
                result = self.check_keyword_ranking(keyword_obj.keyword, project.product_id)
                results.append(result)
                if result['success']:
                    success_count += 1
            
            return {
                'success': success_count > 0,
                'message': f"ìˆœìœ„ í™•ì¸ ì™„ë£Œ: {success_count}/{len(keywords)} í‚¤ì›Œë“œ",
                'results': results,
                'success_count': success_count,
                'total_count': len(keywords)
            }
            
        except Exception as e:
            logger.error(f"í”„ë¡œì íŠ¸ ìˆœìœ„ í™•ì¸ ë¶„ì„ ì‹¤íŒ¨: {e}")
            return {
                'success': False,
                'message': f"ìˆœìœ„ í™•ì¸ ì¤‘ ì˜¤ë¥˜ ë°œìƒ: {e}",
                'results': [],
                'success_count': 0,
                'total_count': len(keywords) if keywords else 0
            }
    
    def process_single_keyword_ranking(self, keyword_obj, product_id: str) -> Tuple[Any, bool]:
        """ë‹¨ì¼ í‚¤ì›Œë“œ ìˆœìœ„ í™•ì¸ ì²˜ë¦¬ (engine_localì—ì„œ ì´ë™)"""
        try:
            # ìˆœìœ„ í™•ì¸
            result = self.check_keyword_ranking(keyword_obj.keyword, product_id)
            logger.info(f"ìˆœìœ„ í™•ì¸ ê²°ê³¼: {keyword_obj.keyword} -> ìˆœìœ„: {result['rank']}, ì„±ê³µ: {result['success']}")
            return result, True
            
        except Exception as e:
            logger.error(f"í‚¤ì›Œë“œ {keyword_obj.keyword} ìˆœìœ„ í™•ì¸ ì‹¤íŒ¨: {e}")
            failed_result = self._create_failed_ranking_result(keyword_obj.keyword, str(e))
            return failed_result, False
    
    def _create_failed_ranking_result(self, keyword: str, error: str) -> RankingCheckDTO:
        """ì‹¤íŒ¨í•œ ìˆœìœ„ ê²°ê³¼ ìƒì„±"""
        return RankingCheckDTO(
            keyword=keyword,
            success=False,
            rank=RANK_OUT_OF_RANGE,
            error=error
        )
    
    def analyze_and_add_keyword(self, keyword: str) -> Dict[str, Any]:
        """í‚¤ì›Œë“œ ë¶„ì„ ë° ì¶”ê°€ ë¡œì§ (engine_localì—ì„œ ì´ë™)"""
        try:
            # í‚¤ì›Œë“œ ë¶„ì„ ìˆ˜í–‰
            analysis = self.analyze_keyword_for_tracking(keyword)
            
            if analysis['success']:
                return {
                    'success': True,
                    'keyword': keyword,
                    'category': analysis['category'],
                    'monthly_volume': analysis['monthly_volume'],
                    'ready_for_db': True
                }
            else:
                return {
                    'success': False,
                    'keyword': keyword,
                    'category': '-',
                    'monthly_volume': -1,
                    'error_message': analysis.get('error_message', 'ë¶„ì„ ì‹¤íŒ¨'),
                    'ready_for_db': False
                }
                
        except Exception as e:
            logger.error(f"í‚¤ì›Œë“œ '{keyword}' ë¶„ì„/ì¶”ê°€ ë¡œì§ ì‹¤íŒ¨: {e}")
            return {
                'success': False,
                'keyword': keyword,
                'category': '-',
                'monthly_volume': -1,
                'error_message': str(e),
                'ready_for_db': False
            }
    
    def refresh_product_info_analysis(self, project) -> Dict[str, Any]:
        """í”„ë¡œì íŠ¸ ìƒí’ˆ ì •ë³´ ìƒˆë¡œê³ ì¹¨ ë¶„ì„ (engine_localì—ì„œ ì´ë™)"""
        try:
            logger.info(f"í”„ë¡œì íŠ¸ ì •ë³´ ìƒˆë¡œê³ ì¹¨ ì‹œì‘: {project.current_name}")
            
            # ìƒí’ˆ ì •ë³´ ì¬ì¡°íšŒ
            product_info_dict = smart_product_search(project.current_name, project.product_id)
            
            if not product_info_dict:
                return {
                    'success': False,
                    'message': f'{project.current_name} ìƒí’ˆ ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.',
                    'changes': []
                }
            
            # ë³€ê²½ì‚¬í•­ ë¶„ì„ì€ engineì— ìœ„ì„
            from .engine_local import rank_tracking_engine
            from .models import ProductInfo
            
            # ë³€ê²½ì‚¬í•­ ë¶„ì„
            new_info = {
                'current_name': clean_product_name(product_info_dict.get('name', '')),
                'price': product_info_dict.get('price', 0),
                'category': product_info_dict.get('category', ''),
                'store_name': product_info_dict.get('store_name', ''),
            }
            
            # ë³€ê²½ì‚¬í•­ ê°ì§€ëŠ” ì—”ì§„ì— ìœ„ì„ (ìˆœìˆ˜ ê³„ì‚°)
            changes_detected = rank_tracking_engine.detect_project_changes(project, new_info)
            
            # ProductInfo ê°ì²´ ìƒì„±
            product_info = ProductInfo(
                product_id=product_info_dict.get('product_id', ''),
                name=new_info['current_name'],
                price=new_info['price'],
                category=new_info['category'],
                store_name=new_info['store_name'],
                description=product_info_dict.get('description', ''),
                image_url=product_info_dict.get('image_url', ''),
                url=product_info_dict.get('url', '')
            )
            
            return {
                'success': True,
                'new_info': new_info,
                'changes': changes_detected,
                'product_info': product_info
            }
            
        except Exception as e:
            logger.error(f"í”„ë¡œì íŠ¸ ì •ë³´ ìƒˆë¡œê³ ì¹¨ ë¶„ì„ ì‹¤íŒ¨: {e}")
            return {
                'success': False,
                'message': f'ìƒí’ˆ ì •ë³´ ìƒˆë¡œê³ ì¹¨ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}',
                'changes': []
            }
    
    def get_product_category_analysis(self, product_id: str) -> Dict[str, Any]:
        """ìƒí’ˆ ì¹´í…Œê³ ë¦¬ ì¡°íšŒ ë¶„ì„ (engine_localì—ì„œ ì´ë™)"""
        try:
            # smart_product_searchë¥¼ í†µí•´ ìƒí’ˆ ì •ë³´ ì¡°íšŒ
            product_info = smart_product_search(f"ìƒí’ˆID_{product_id}", product_id)
            if product_info and 'category' in product_info:
                return {
                    'success': True,
                    'category': product_info['category'],
                    'product_id': product_id
                }
            
            return {
                'success': False,
                'category': '',
                'product_id': product_id,
                'error_message': 'ìƒí’ˆ ì •ë³´ ì¡°íšŒ ì‹¤íŒ¨'
            }
            
        except Exception as e:
            logger.error(f"ìƒí’ˆ ì¹´í…Œê³ ë¦¬ ë¶„ì„ ì‹¤íŒ¨: {e}")
            return {
                'success': False,
                'category': '',
                'product_id': product_id,
                'error_message': str(e)
            }
    
    def process_keyword_info_analysis(self, keyword: str) -> Dict[str, Any]:
        """í‚¤ì›Œë“œ ì •ë³´ ë¶„ì„ ì²˜ë¦¬ (engine_localì—ì„œ ì´ë™)"""
        try:
            # ì¹´í…Œê³ ë¦¬ì™€ ì›”ê²€ìƒ‰ëŸ‰ ì¡°íšŒ
            category = self.get_keyword_category(keyword)
            monthly_volume = self.get_keyword_monthly_volume(keyword)
            
            return {
                'success': True,
                'category': category if category else '-',
                'monthly_volume': monthly_volume if monthly_volume is not None else -1,
                'keyword': keyword
            }
            
        except Exception as e:
            logger.error(f"í‚¤ì›Œë“œ ì •ë³´ ë¶„ì„ ì‹¤íŒ¨: {keyword}: {e}")
            return {
                'success': False,
                'category': '-',
                'monthly_volume': -1,
                'keyword': keyword,
                'error_message': str(e)
            }


def to_export_row(keyword_row: dict) -> dict:
    """í‚¤ì›Œë“œ í–‰ ë°ì´í„°ë¥¼ ì—‘ì…€ ë‚´ë³´ë‚´ê¸°ìš© ë·°ëª¨ë¸ë¡œ ë³€í™˜"""
    return {
        "í‚¤ì›Œë“œ": keyword_row.get("keyword", ""),
        "ì¹´í…Œê³ ë¦¬": keyword_row.get("category", "") or "-",
        "ì›”ê²€ìƒ‰ëŸ‰": format_monthly_volume(keyword_row.get("monthly_volume")),
        "í˜„ì¬ìˆœìœ„": format_rank(keyword_row.get("rank_position")),
        "ì ê²€ì‹œê°": format_datetime(keyword_row.get("search_date")),
    }


def to_display_row(keyword_row: dict) -> dict:
    """í‚¤ì›Œë“œ í–‰ ë°ì´í„°ë¥¼ UI í‘œì‹œìš© ë·°ëª¨ë¸ë¡œ ë³€í™˜"""
    return {
        "keyword": keyword_row.get("keyword", ""),
        "category": keyword_row.get("category", "") or "-",
        "monthly_volume_display": format_monthly_volume(keyword_row.get("monthly_volume")),
        "rank_display": format_rank(keyword_row.get("rank_position")),
        "search_date_display": format_datetime(keyword_row.get("search_date")),
        # ì›ë³¸ ë°ì´í„°ë„ ìœ ì§€ (ì •ë ¬/í•„í„°ë§ìš©)
        "monthly_volume_raw": keyword_row.get("monthly_volume"),
        "rank_position_raw": keyword_row.get("rank_position"),
        "search_date_raw": keyword_row.get("search_date"),
    }


class RankTrackingExcelExporter:
    """ìˆœìœ„ ì¶”ì  Excel ë‚´ë³´ë‚´ê¸° ì–´ëŒ‘í„° (íŒŒì¼ I/O ë‹´ë‹¹)"""
    
    def get_default_filename(self, project_id: Optional[int] = None) -> str:
        """ê¸°ë³¸ íŒŒì¼ëª… ìƒì„±"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if project_id:
            try:
                from .service import rank_tracking_service
                project = rank_tracking_service.get_project_by_id(project_id)
                if project:
                    # íŒŒì¼ëª…ì— ì‚¬ìš©í•  ìˆ˜ ì—†ëŠ” ë¬¸ì ì œê±°
                    safe_name = project.current_name.replace('/', '_').replace('\\', '_')
                    safe_name = safe_name.replace(':', '_').replace('*', '_')
                    safe_name = safe_name.replace('?', '_').replace('"', '_')
                    safe_name = safe_name.replace('<', '_').replace('>', '_')
                    safe_name = safe_name.replace('|', '_')
                    
                    return f"ìˆœìœ„ì´ë ¥_{safe_name}_{timestamp}.xlsx"
            except Exception as e:
                logger.error(f"í”„ë¡œì íŠ¸ëª… ì¡°íšŒ ì‹¤íŒ¨: {e}")
        
        return f"ìˆœìœ„ì´ë ¥_ë°ì´í„°_{timestamp}.xlsx"
    
    def export_ranking_history_to_excel(self, project_id: int, file_path: str, ranking_data: list = None) -> bool:
        """ìˆœìœ„ ì´ë ¥ì„ Excelë¡œ ë‚´ë³´ë‚´ê¸° (ì›ë³¸ê³¼ ë™ì¼)"""
        try:
            from .service import rank_tracking_service
            
            # í”„ë¡œì íŠ¸ ì •ë³´ ì¡°íšŒ
            project = rank_tracking_service.get_project_by_id(project_id)
            if not project:
                logger.error(f"í”„ë¡œì íŠ¸ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {project_id}")
                return False
            
            # í‚¤ì›Œë“œ ì •ë³´ ì¡°íšŒ
            keywords = rank_tracking_service.get_project_keywords(project_id)
            if not keywords:
                logger.error("í‚¤ì›Œë“œê°€ ì—†ìŠµë‹ˆë‹¤")
                return False
            
            # í”„ë¡œì íŠ¸ ìˆœìœ„ ê°œìš” ë°ì´í„° ì‚¬ìš©
            overview = rank_tracking_service.get_project_overview(project_id)
            all_dates = overview.get('dates', [])[:10]  # ìµœëŒ€ 10ê°œ ë‚ ì§œ
            keywords_data = overview.get('keywords', {})
            
            logger.info(f"ë””ë²„ê¹…: í‚¤ì›Œë“œ ìˆ˜ = {len(keywords)}")
            logger.info(f"ë””ë²„ê¹…: ì „ì²´ ë‚ ì§œ ìˆ˜ = {len(all_dates)}, ë‚ ì§œë“¤ = {all_dates}")
            
            # í‚¤ì›Œë“œë³„ ìˆœìœ„ ë°ì´í„° êµ¬ì„±
            keyword_ranking_data = []
            for keyword_obj in keywords:
                # keywords_dataëŠ” {keyword: {date: rank}} êµ¬ì¡°
                rankings = keywords_data.get(keyword_obj.keyword, {})
                
                logger.info(f"ë””ë²„ê¹…: í‚¤ì›Œë“œ '{keyword_obj.keyword}' ìˆœìœ„ ì´ë ¥ ìˆ˜ = {len(rankings)}")
                
                # ë‚ ì§œë³„ ìˆœìœ„ ë§¤í•‘ (overview ë°ì´í„° í˜•ì‹ì— ë§ì¶¤)
                rank_by_date = {}
                for date in all_dates:
                    if date in rankings:
                        # rankings[date]ëŠ” ì§ì ‘ rank ê°’ (ì •ìˆ˜)
                        rank = rankings[date]
                        rank_by_date[date] = rank
                        logger.info(f"ë””ë²„ê¹…: ìˆœìœ„ ë°ì´í„° - í‚¤ì›Œë“œ: {keyword_obj.keyword}, ë‚ ì§œ: {date}, ìˆœìœ„: {rank}")
                
                keyword_ranking_data.append({
                    'keyword': keyword_obj.keyword,
                    'category': keyword_obj.category or '-',
                    'monthly_volume': keyword_obj.monthly_volume if keyword_obj.monthly_volume is not None else -1,
                    'rank_by_date': rank_by_date
                })
            
            # ë‚ ì§œ ì •ë ¬ (ìµœì‹ ìˆœ) ë° í˜•ì‹ ë³€í™˜
            sorted_dates = []
            formatted_dates = []
            for date in all_dates:
                # ë‚ ì§œë¥¼ MM/DD HH:MM í˜•ì‹ìœ¼ë¡œ ë³€í™˜
                try:
                    if isinstance(date, str):
                        dt = _to_dt(date)
                        if dt:
                            formatted_date = dt.strftime("%m/%d %H:%M")
                            sorted_dates.append(date)  # ì›ë³¸ ë‚ ì§œ (í‚¤ ë§¤ì¹­ìš©)
                            formatted_dates.append(formatted_date)  # í‘œì‹œìš© ë‚ ì§œ
                            logger.info(f"ë””ë²„ê¹…: ë‚ ì§œ ë³€í™˜ - {date} -> {formatted_date}")
                        else:
                            logger.warning(f"ë””ë²„ê¹…: ë‚ ì§œ íŒŒì‹± ì‹¤íŒ¨ - {date}")
                            continue
                except Exception as e:
                    logger.warning(f"ë””ë²„ê¹…: ë‚ ì§œ ë³€í™˜ ì‹¤íŒ¨ - {date}: {e}")
                    continue
            
            logger.info(f"ë””ë²„ê¹…: ìµœì¢… ë‚ ì§œ ìˆ˜ = {len(sorted_dates)}, ë³€í™˜ëœ ë‚ ì§œë“¤ = {formatted_dates}")
            
            # ì—‘ì…€ ë°ì´í„° êµ¬ì„±
            excel_data = []
            
            # 1. ê¸°ë³¸ì •ë³´ ì„¹ì…˜ (ì‚¬ì§„ê³¼ ë˜‘ê°™ì´)
            excel_data.extend([
                [f"ğŸ“Š {project.current_name}", "", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["ìƒí’ˆ ID", project.product_id, "", "", "", "", "", "", ""],
                ["ìƒí’ˆëª…", project.current_name, "", "", "", "", "", "", ""],
                ["ìŠ¤í† ì–´ëª…", project.store_name or "-", "", "", "", "", "", "", ""],
                ["ê°€ê²©", format_price_krw(project.price), "", "", "", "", "", "", ""],
                ["ì¹´í…Œê³ ë¦¬", project.category or "-", "", "", "", "", "", "", ""],
                ["ë“±ë¡ì¼", self._format_date(project.created_at) if project.created_at else "-", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["ğŸ” í‚¤ì›Œë“œ ìˆœìœ„ í˜„í™©", "", "", "", "", "", "", "", ""]
            ])
            
            # 2. í‚¤ì›Œë“œ ìˆœìœ„ í…Œì´ë¸” í—¤ë” (í‘œì‹œìš© ë‚ ì§œ ì‚¬ìš©)
            header_row = ["í‚¤ì›Œë“œ", "ì¹´í…Œê³ ë¦¬", "ì›”ê²€ìƒ‰ëŸ‰"]
            header_row.extend(formatted_dates)  # ë³€í™˜ëœ ë‚ ì§œë“¤ ì¶”ê°€
            excel_data.append(header_row)
            
            # 3. í‚¤ì›Œë“œë³„ ìˆœìœ„ ë°ì´í„°
            for kw_data in keyword_ranking_data:
                # ìƒˆë¡œìš´ í¬ë§·í„° ì‚¬ìš©
                volume_display = format_monthly_volume(kw_data['monthly_volume'])
                
                data_row = [
                    kw_data['keyword'],
                    kw_data['category'],
                    volume_display
                ]
                
                # ê° ë‚ ì§œë³„ ìˆœìœ„ ì¶”ê°€ (ì›ë³¸ ë‚ ì§œë¡œ í‚¤ ë§¤ì¹­)
                for date in sorted_dates:
                    rank = kw_data['rank_by_date'].get(date, "")
                    if rank:
                        if rank == 999 or rank > 200:  # RANK_OUT_OF_RANGE ë˜ëŠ” 200ìœ„ ì´ˆê³¼
                            data_row.append("200+")
                        else:
                            data_row.append(f"{rank}ìœ„")
                    else:
                        data_row.append("")
                
                excel_data.append(data_row)
            
            # ì—‘ì…€ íŒŒì¼ ìƒì„± (í†µí•©ëœ ë°©ì‹ ì‚¬ìš©)
            success = self._create_single_excel_file(file_path, excel_data)
            
            if success:
                logger.info(f"ìˆœìœ„ ì´ë ¥ ì—‘ì…€ íŒŒì¼ ìƒì„± ì™„ë£Œ: {file_path}")
            else:
                logger.error("ìˆœìœ„ ì´ë ¥ ì—‘ì…€ íŒŒì¼ ìƒì„± ì‹¤íŒ¨")
            
            return success
            
        except Exception as e:
            logger.error(f"ìˆœìœ„ ì´ë ¥ ì—‘ì…€ ì €ì¥ ì¤‘ ì˜¤ë¥˜: {e}")
            return False
    
    def export_multiple_projects_to_excel(self, project_ids: List[int], file_path: str) -> bool:
        """ì—¬ëŸ¬ í”„ë¡œì íŠ¸ë¥¼ ì—‘ì…€ë¡œ ì €ì¥ (ê° í”„ë¡œì íŠ¸ë³„ ì™„ì „í•œ ìŠ¤íƒ€ì¼ ì ìš©)"""
        try:
            logger.info(f"ë‹¤ì¤‘ í”„ë¡œì íŠ¸ ì—‘ì…€ ì €ì¥ ì‹œì‘: {len(project_ids)}ê°œ í”„ë¡œì íŠ¸")
            
            if not project_ids:
                logger.error("í”„ë¡œì íŠ¸ IDê°€ ì—†ìŠµë‹ˆë‹¤")
                return False
            
            import openpyxl
            from .service import rank_tracking_service
            
            # ìƒˆ ì›Œí¬ë¶ ìƒì„±
            workbook = openpyxl.Workbook()
            # ê¸°ë³¸ ì‹œíŠ¸ ì œê±°
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # ê° í”„ë¡œì íŠ¸ë³„ë¡œ ì‹œíŠ¸ ìƒì„±
            for i, project_id in enumerate(project_ids, 1):
                try:
                    logger.info(f"í”„ë¡œì íŠ¸ {project_id} ì²˜ë¦¬ ì¤‘... ({i}/{len(project_ids)})")
                    
                    # í”„ë¡œì íŠ¸ ì •ë³´ ì¡°íšŒ
                    project = rank_tracking_service.get_project_by_id(project_id)
                    if not project:
                        logger.warning(f"í”„ë¡œì íŠ¸ {project_id}ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŒ")
                        continue
                    
                    # í”„ë¡œì íŠ¸ë³„ ì—‘ì…€ ë°ì´í„° ìƒì„± (ê¸°ì¡´ ë¡œì§ ì¬ì‚¬ìš©)
                    excel_data = self._generate_project_excel_data(project_id)
                    if not excel_data:
                        logger.warning(f"í”„ë¡œì íŠ¸ {project_id} ë°ì´í„° ìƒì„± ì‹¤íŒ¨")
                        continue
                    
                    # ì‹œíŠ¸ ì´ë¦„ì„ ìƒí’ˆëª…ìœ¼ë¡œ ì„¤ì • (ì—‘ì…€ ì‹œíŠ¸ëª… ì œí•œ ê³ ë ¤)
                    sheet_name = project.current_name[:31] if project.current_name else f"í”„ë¡œì íŠ¸{i}"
                    # ì—‘ì…€ ì‹œíŠ¸ëª…ì— ì‚¬ìš©í•  ìˆ˜ ì—†ëŠ” ë¬¸ì ì œê±°
                    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
                    for char in invalid_chars:
                        sheet_name = sheet_name.replace(char, '_')
                    
                    # ìƒˆ ì‹œíŠ¸ ìƒì„±
                    new_sheet = workbook.create_sheet(title=sheet_name)
                    
                    # ì™„ì „í•œ ìŠ¤íƒ€ì¼ì´ ì ìš©ëœ ë°ì´í„° ì…ë ¥ (ê¸°ì¡´ _create_excel_file ë¡œì§ ì‚¬ìš©)
                    self._apply_excel_data_to_sheet(new_sheet, excel_data)
                    
                    logger.info(f"í”„ë¡œì íŠ¸ {project_id} ì‹œíŠ¸ '{sheet_name}' ìƒì„± ì™„ë£Œ")
                    
                except Exception as e:
                    logger.error(f"í”„ë¡œì íŠ¸ {project_id} ì²˜ë¦¬ ì‹¤íŒ¨: {e}")
                    continue
            
            if len(workbook.sheetnames) == 0:
                logger.error("ìƒì„±ëœ ì‹œíŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤")
                return False
            
            # íŒŒì¼ ì €ì¥
            workbook.save(file_path)
            workbook.close()
            
            logger.info(f"ë‹¤ì¤‘ í”„ë¡œì íŠ¸ ì—‘ì…€ ì €ì¥ ì™„ë£Œ: ì´ {len(workbook.sheetnames)}ê°œ ì‹œíŠ¸")
            return True
            
        except Exception as e:
            logger.error(f"ë‹¤ì¤‘ í”„ë¡œì íŠ¸ ì—‘ì…€ ì €ì¥ ì¤‘ ì¹˜ëª…ì  ì˜¤ë¥˜: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def _generate_project_excel_data(self, project_id: int) -> list:
        """í”„ë¡œì íŠ¸ë³„ ì—‘ì…€ ë°ì´í„° ìƒì„± (ê¸°ì¡´ export_ranking_history_to_excel ë¡œì§ ì¬ì‚¬ìš©)"""
        try:
            from .service import rank_tracking_service
            
            # í”„ë¡œì íŠ¸ ì •ë³´ ì¡°íšŒ
            project = rank_tracking_service.get_project_by_id(project_id)
            if not project:
                logger.error(f"í”„ë¡œì íŠ¸ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {project_id}")
                return []
            
            # í‚¤ì›Œë“œ ì •ë³´ ì¡°íšŒ
            keywords = rank_tracking_service.get_project_keywords(project_id)
            if not keywords:
                logger.error("í‚¤ì›Œë“œê°€ ì—†ìŠµë‹ˆë‹¤")
                return []
            
            # í”„ë¡œì íŠ¸ ìˆœìœ„ ê°œìš” ë°ì´í„° ì‚¬ìš©
            overview = rank_tracking_service.get_project_overview(project_id)
            all_dates = overview.get('dates', [])[:10]  # ìµœëŒ€ 10ê°œ ë‚ ì§œ
            keywords_data = overview.get('keywords', {})
            
            # í‚¤ì›Œë“œë³„ ìˆœìœ„ ë°ì´í„° êµ¬ì„±
            keyword_ranking_data = []
            for keyword_obj in keywords:
                rankings = keywords_data.get(keyword_obj.keyword, {})
                rank_by_date = {}
                for date in all_dates:
                    if date in rankings:
                        rank = rankings[date]
                        rank_by_date[date] = rank
                
                keyword_ranking_data.append({
                    'keyword': keyword_obj.keyword,
                    'category': keyword_obj.category or '-',
                    'monthly_volume': keyword_obj.monthly_volume if keyword_obj.monthly_volume is not None else -1,
                    'rank_by_date': rank_by_date
                })
            
            # ë‚ ì§œ ì •ë ¬ ë° í˜•ì‹ ë³€í™˜
            sorted_dates = []
            formatted_dates = []
            for date in all_dates:
                try:
                    if isinstance(date, str):
                        dt = _to_dt(date)
                        if dt:
                            formatted_date = dt.strftime("%m/%d %H:%M")
                            sorted_dates.append(date)
                            formatted_dates.append(formatted_date)
                except Exception as e:
                    continue
            
            # ì—‘ì…€ ë°ì´í„° êµ¬ì„±
            excel_data = []
            
            # 1. ê¸°ë³¸ì •ë³´ ì„¹ì…˜
            excel_data.extend([
                [f"ğŸ“Š {project.current_name}", "", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["ìƒí’ˆ ID", project.product_id, "", "", "", "", "", "", ""],
                ["ìƒí’ˆëª…", project.current_name, "", "", "", "", "", "", ""],
                ["ìŠ¤í† ì–´ëª…", project.store_name or "-", "", "", "", "", "", "", ""],
                ["ê°€ê²©", format_price_krw(project.price), "", "", "", "", "", "", ""],
                ["ì¹´í…Œê³ ë¦¬", project.category or "-", "", "", "", "", "", "", ""],
                ["ë“±ë¡ì¼", self._format_date(project.created_at) if project.created_at else "-", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["", "", "", "", "", "", "", "", ""],
                ["ğŸ” í‚¤ì›Œë“œ ìˆœìœ„ í˜„í™©", "", "", "", "", "", "", "", ""]
            ])
            
            # 2. í‚¤ì›Œë“œ ìˆœìœ„ í…Œì´ë¸” í—¤ë”
            header_row = ["í‚¤ì›Œë“œ", "ì¹´í…Œê³ ë¦¬", "ì›”ê²€ìƒ‰ëŸ‰"]
            header_row.extend(formatted_dates)
            excel_data.append(header_row)
            
            # 3. í‚¤ì›Œë“œë³„ ìˆœìœ„ ë°ì´í„°
            for kw_data in keyword_ranking_data:
                volume_display = format_monthly_volume(kw_data['monthly_volume'])
                
                data_row = [
                    kw_data['keyword'],
                    kw_data['category'],
                    volume_display
                ]
                
                # ê° ë‚ ì§œë³„ ìˆœìœ„ ì¶”ê°€
                for date in sorted_dates:
                    rank = kw_data['rank_by_date'].get(date, "")
                    if rank:
                        if rank == 999 or rank > 200:
                            data_row.append("200+")
                        else:
                            data_row.append(f"{rank}ìœ„")
                    else:
                        data_row.append("")
                
                excel_data.append(data_row)
            
            return excel_data
            
        except Exception as e:
            logger.error(f"í”„ë¡œì íŠ¸ ì—‘ì…€ ë°ì´í„° ìƒì„± ì¤‘ ì˜¤ë¥˜: {e}")
            return []
    
    def _apply_excel_data_to_sheet(self, worksheet, excel_data: list):
        """ì›Œí¬ì‹œíŠ¸ì— ì—‘ì…€ ë°ì´í„°ì™€ ìŠ¤íƒ€ì¼ ì ìš© (ê¸°ì¡´ _create_excel_file ë¡œì§ ì¬ì‚¬ìš©)"""
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # ë°ì´í„° ì…ë ¥ ë° ìŠ¤íƒ€ì¼ ì ìš©
            for row_idx, row_data in enumerate(excel_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    # ì›”ê²€ìƒ‰ëŸ‰ê³¼ ìˆœìœ„ ì»¬ëŸ¼ì€ ìˆ«ìë¡œ ì €ì¥í•˜ì—¬ ì •ë ¬ ê°€ëŠ¥í•˜ê²Œ í•¨
                    if row_idx > 12 and col_idx == 3:  # ì›”ê²€ìƒ‰ëŸ‰ ì»¬ëŸ¼
                        try:
                            if isinstance(cell_value, str):
                                if cell_value == "ë¯¸ìˆ˜ì§‘" or cell_value == "N/A":
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                                elif cell_value == "0":
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=0)
                                    cell.number_format = '#,##0'
                                elif cell_value.replace(',', '').isdigit():
                                    numeric_value = int(cell_value.replace(',', ''))
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=numeric_value)
                                    cell.number_format = '#,##0'
                                else:
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                            elif isinstance(cell_value, int):
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                                cell.number_format = '#,##0'
                            else:
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        except:
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    elif row_idx > 12 and col_idx > 3:  # ìˆœìœ„ ì»¬ëŸ¼ë“¤
                        try:
                            if isinstance(cell_value, str):
                                if "200+" in cell_value or "200ìœ„ë°–" in cell_value:
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=999)  # 200ìœ„ë°–ì€ 999ë¡œ ì €ì¥
                                    cell.number_format = '"200ìœ„ë°–"'
                                elif "ìœ„" in cell_value:
                                    rank_num = int(cell_value.replace("ìœ„", ""))
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=rank_num)
                                    cell.number_format = '0"ìœ„"'
                                else:
                                    cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                            else:
                                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        except:
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    else:
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    # ìŠ¤íƒ€ì¼ ì ìš©
                    if row_idx == 1:  # ì œëª© í–‰
                        cell.font = Font(size=14, bold=True)
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.font = Font(color="FFFFFF", size=14, bold=True)
                    elif row_idx == 11:  # í‚¤ì›Œë“œ ìˆœìœ„ í˜„í™© í—¤ë”
                        cell.font = Font(size=12, bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    elif row_idx == 12:  # í…Œì´ë¸” í—¤ë”
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif row_idx > 12:  # ë°ì´í„° í–‰
                        if col_idx <= 3:  # í‚¤ì›Œë“œ, ì¹´í…Œê³ ë¦¬, ì›”ê²€ìƒ‰ëŸ‰ ì»¬ëŸ¼
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:  # ìˆœìœ„ ì»¬ëŸ¼ë“¤
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            # ìˆœìœ„ì— ë”°ë¥¸ ìƒ‰ìƒ ì ìš© (UIì™€ ë™ì¼í•œ ê¸°ì¤€)
                            if isinstance(cell.value, (int, float)):
                                rank_num = int(cell.value)
                                if rank_num <= 10:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # ì—°í•œ ì´ˆë¡ìƒ‰ (1-10ìœ„)
                                elif rank_num <= 40:
                                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # ì—°í•œ ë…¸ë€ìƒ‰ (11-40ìœ„)
                                elif rank_num <= 200:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # ì—°í•œ ë¹¨ê°„ìƒ‰ (41-200ìœ„)
                                elif rank_num == 999 or rank_num > 200:  # 200ìœ„ë°– (999 ë˜ëŠ” 201 ì´ìƒ)
                                    cell.fill = PatternFill(start_color="D63384", end_color="D63384", fill_type="solid")  # ë” ì§„í•œ ë¹¨ê°„ìƒ‰ (200ìœ„ë°–)
                                    cell.font = Font(color="FFFFFF")  # í°ìƒ‰ í…ìŠ¤íŠ¸ë¡œ ê°€ë…ì„± í–¥ìƒ
            
            # ì»¬ëŸ¼ ë„ˆë¹„ ì„¤ì •
            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                
                if col_idx == 1:  # í‚¤ì›Œë“œ ì»¬ëŸ¼
                    worksheet.column_dimensions[column_letter].width = 20
                elif col_idx == 2:  # ì¹´í…Œê³ ë¦¬ ì»¬ëŸ¼
                    worksheet.column_dimensions[column_letter].width = 30
                elif col_idx == 3:  # ì›”ê²€ìƒ‰ëŸ‰ ì»¬ëŸ¼
                    worksheet.column_dimensions[column_letter].width = 12
                else:  # ìˆœìœ„ ì»¬ëŸ¼ë“¤
                    worksheet.column_dimensions[column_letter].width = 15
                    
        except Exception as e:
            logger.error(f"ì›Œí¬ì‹œíŠ¸ ìŠ¤íƒ€ì¼ ì ìš© ì¤‘ ì˜¤ë¥˜: {e}")

    def _format_date(self, date_value):
        """ë‚ ì§œ í˜•ì‹ì„ ì•ˆì „í•˜ê²Œ ë³€í™˜"""
        try:
            if isinstance(date_value, str):
                # ë¬¸ìì—´ì¸ ê²½ìš° datetimeìœ¼ë¡œ ë³€í™˜
                dt = _to_dt(date_value)
                return dt.strftime("%Y-%m-%d") if dt else str(date_value)
            elif hasattr(date_value, 'strftime'):
                # datetime ê°ì²´ì¸ ê²½ìš°
                return date_value.strftime("%Y-%m-%d")
            else:
                # ê¸°íƒ€ ê²½ìš°ëŠ” ë¬¸ìì—´ë¡œ ë³€í™˜
                return str(date_value)
        except Exception as e:
            logger.warning(f"ë‚ ì§œ í˜•ì‹ ë³€í™˜ ì‹¤íŒ¨: {e}")
            return str(date_value) if date_value else "-"
    
    
    def _create_single_excel_file(self, file_path: str, excel_data: list) -> bool:
        """ë‹¨ì¼ ì—‘ì…€ íŒŒì¼ ìƒì„± (í†µí•©ëœ ë°©ì‹ ì‚¬ìš©)"""
        try:
            import openpyxl
            
            # ì›Œí¬ë¶ ìƒì„±
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "ìˆœìœ„ì´ë ¥"
            
            # ìŠ¤íƒ€ì¼ ì ìš© (ê³µìš© ë©”ì„œë“œ ì‚¬ìš©)
            self._apply_excel_data_to_sheet(worksheet, excel_data)
            
            # íŒŒì¼ ì €ì¥
            workbook.save(file_path)
            return True
            
        except Exception as e:
            logger.error(f"ì—‘ì…€ íŒŒì¼ ìƒì„± ì¤‘ ì˜¤ë¥˜: {e}")
            return False


# ì „ì—­ ì–´ëŒ‘í„° ì¸ìŠ¤í„´ìŠ¤
rank_tracking_adapter = RankTrackingAdapter()
rank_tracking_excel_exporter = RankTrackingExcelExporter()