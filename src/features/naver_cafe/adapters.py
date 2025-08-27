"""
네이버 카페 DB 추출기 어댑터
CLAUDE.md 구조 준수: vendors 호출, 정규화, 파일 I/O 담당
"""
from typing import List, Optional, Dict, Tuple
import time
import csv

# 벤더/파일 I/O만 담당, 비즈니스 로직은 service에서
from src.foundation.logging import get_logger
from .models import CafeInfo, BoardInfo, ExtractedUser
# 실제 사용되는 설정을 adapters에 직접 정의 (CLAUDE.md: 간소화)
RATE_LIMIT_REQUESTS_PER_MINUTE = 30
META_CSV_DOMAINS = ["@naver.com", "@gmail.com", "@daum.net"]

logger = get_logger("features.naver_cafe.adapters")


class NaverCafeDataAdapter:
    """네이버 카페 데이터 어댑터"""
    
    def __init__(self):
        self.rate_limiter = RateLimiter(
            requests_per_minute=RATE_LIMIT_REQUESTS_PER_MINUTE
        )
    
    
    async def search_cafes_by_name(self, query: str, browser_context=None) -> List[CafeInfo]:
        """카페명으로 카페 검색 - 원본과 동일"""
        try:
            await self.rate_limiter.wait_async()
            logger.info(f"카페 검색: {query}")
            
            # URL인지 확인
            if "cafe.naver.com" in query:
                # URL인 경우 직접 접근
                cafe_info = await self.get_cafe_info_by_url(query, browser_context)
                return [cafe_info] if cafe_info else []
            else:
                # 키워드인 경우 검색
                if browser_context:
                    return await self._search_cafes_with_playwright(query, browser_context)
                else:
                    return []
            
        except Exception as e:
            logger.error(f"카페 검색 실패: {e}")
            return []
    
    async def _search_cafes_with_playwright(self, query: str, browser_context) -> List[CafeInfo]:
        """Playwright를 사용한 실제 카페 검색 - 완전 비동기"""
        cafes = []
        page = None
        max_retries = 3
        
        # 브라우저 컨텍스트 확인
        logger.info(f"브라우저 컨텍스트 상태: {browser_context}")
        logger.info(f"브라우저 컨텍스트 타입: {type(browser_context)}")
        
        for attempt in range(max_retries):
            try:
                # 기존 페이지가 있는지 확인하고 재사용
                if (hasattr(self, '_search_page') and self._search_page):
                    page = self._search_page
                else:
                    # 새 페이지 생성하고 저장
                    page = await browser_context.new_page()
                    self._search_page = page
                
                # 네이버 카페 검색 페이지로 이동 (원본과 동일)
                search_url = f"https://section.cafe.naver.com/ca-fe/home/search/cafes?q={query}&od=2"
                logger.info(f"카페 검색 URL: {search_url}")
                await page.goto(search_url, wait_until='networkidle', timeout=20000)
                
                # 카페 목록 대기 (원본과 동일)
                await page.wait_for_selector('.CafeItem', timeout=10000)
                
                # 카페 정보 추출 (원본과 동일)
                cafe_elements = await page.query_selector_all('.CafeItem')
                
                for element in cafe_elements:
                    try:
                        # 카페 이름 (원본과 동일)
                        name_element = await element.query_selector('.cafe_name')
                        cafe_name = await name_element.inner_text() if name_element else ""
                        
                        # 카페 링크 (원본과 동일)
                        link_element = await element.query_selector('a')
                        cafe_url = await link_element.get_attribute('href') if link_element else ""
                        
                        # 회원수 (원본과 동일)
                        member_element = await element.query_selector('.member')
                        member_count = await member_element.inner_text() if member_element else "0"
                        
                        # 카페 ID 추출 (원본과 동일)
                        cafe_id = self._extract_cafe_id_from_url(cafe_url)
                        
                        if cafe_name and cafe_url:
                            cafe_info = CafeInfo(
                                name=cafe_name,
                                url=cafe_url,
                                member_count=member_count,
                                cafe_id=cafe_id,
                                description=f"검색 결과: {cafe_name}"
                            )
                            
                            cafes.append(cafe_info)
                            logger.info(f"카페 발견: {cafe_name} -> {cafe_url}")
                            
                    except Exception as e:
                        logger.debug(f"카페 파싱 실패: {e}")
                        continue
                
                # 성공시 루프 탈출
                if cafes or attempt == max_retries - 1:
                    break
                    
            except Exception as e:
                logger.warning(f"카페 검색 시도 {attempt + 1}/{max_retries} 실패: {e}")
                
                if attempt < max_retries - 1:
                    # 지수적 백오프
                    import asyncio
                    wait_time = 2 ** attempt
                    logger.info(f"{wait_time}초 대기 후 재시도...")
                    await asyncio.sleep(wait_time)
                else:
                    logger.error(f"Playwright 카페 검색 실패: {e}")
                    break
        
        return cafes
    
    def _extract_cafe_id_from_url(self, url: str) -> str:
        """URL에서 카페 ID 추출 - 원본과 동일"""
        try:
            if 'cafe.naver.com/' in url:
                return url.split('cafe.naver.com/')[-1].split('/')[0].split('?')[0]
        except:
            pass
        return ""
    
    
    
    
    async def get_cafe_info_by_url(self, url: str, browser_context=None) -> Optional[CafeInfo]:
        """URL로 카페 정보 가져오기 - 원본과 동일"""
        if not browser_context:
            return None
            
        try:
            # 기존 페이지 재사용 또는 새로 생성
            if (hasattr(self, '_info_page') and self._info_page):
                page = self._info_page
            else:
                page = await browser_context.new_page()
                self._info_page = page
            
            # 카페 메인 URL 생성 (게시글 URL이든 메인 URL이든 메인으로 통일)
            cafe_id = self._extract_cafe_id_from_url(url)
            if not cafe_id:
                logger.error(f"올바르지 않은 카페 URL: {url}")
                return None
                
            cafe_main_url = f"https://cafe.naver.com/{cafe_id}"
            await page.goto(cafe_main_url, wait_until="networkidle", timeout=20000)
            
            # 카페 이름 추출
            try:
                cafe_name = await page.locator('h1.d-none').inner_text()
            except:
                cafe_name = f"카페 {cafe_id}"
            
            # 회원수 추출 
            try:
                member_count = await page.locator('li.mem-cnt-info em').inner_text()
            except:
                member_count = "정보 없음"
            
            return CafeInfo(
                name=cafe_name,
                url=cafe_main_url,
                member_count=member_count,
                cafe_id=cafe_id
            )
            
        except Exception as e:
            logger.error(f"URL 카페 정보 추출 실패: {e}")
            return None
    
    
    async def get_cafe_boards(self, cafe_info: CafeInfo, browser_context=None) -> List[BoardInfo]:
        """카페의 게시판 목록 조회 - 완전 비동기"""
        try:
            await self.rate_limiter.wait_async()
            logger.info(f"비동기 게시판 목록 조회: {cafe_info.name}")
            
            if browser_context:
                return await self._get_boards_with_playwright(cafe_info, browser_context)
            else:
                return []
            
        except Exception as e:
            logger.error(f"비동기 게시판 목록 조회 실패: {e}")
            return []
    
    async def _get_boards_with_playwright(self, cafe_info: CafeInfo, browser_context) -> List[BoardInfo]:
        """Playwright를 사용한 실제 게시판 목록 조회 - 원본과 동일"""
        boards = []
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                # 기존 페이지 재사용 또는 새로 생성
                if (hasattr(self, '_board_page') and self._board_page):
                    page = self._board_page
                else:
                    page = await browser_context.new_page()
                    self._board_page = page
                
                # 카페 메인 페이지로 이동 
                logger.info(f"게시판 목록 로딩: {cafe_info.url}")
                await page.goto(cafe_info.url, wait_until='networkidle', timeout=20000)
                
                # 게시판 목록 요소 대기 
                await page.wait_for_selector('ul.cafe-menu-list', timeout=10000)
                
                # 게시판 목록 추출 
                board_elements = await page.query_selector_all('ul.cafe-menu-list li a')
                
                for element in board_elements:
                    try:
                        href = await element.get_attribute('href')
                        if href and '/ArticleList.nhn' in href:
                            board_name = await element.inner_text()
                            board_id = self._extract_board_id_from_url(href)
                            
                            # 상대 경로를 절대 경로로 변환 
                            if href.startswith('/'):
                                full_url = f"https://cafe.naver.com{href}"
                            else:
                                full_url = href
                            
                            board_info = BoardInfo(
                                name=board_name.strip(),
                                url=full_url,
                                board_id=board_id,
                                article_count=0
                            )
                            
                            boards.append(board_info)
                            logger.info(f"게시판 발견: {board_name}")
                            
                    except Exception as e:
                        logger.debug(f"게시판 파싱 실패: {e}")
                        continue
                
                # 성공시 루프 탈출
                if boards or attempt == max_retries - 1:
                    break
                    
            except Exception as e:
                logger.warning(f"게시판 로딩 시도 {attempt + 1}/{max_retries} 실패: {e}")
                
                if attempt < max_retries - 1:
                    # 지수적 백오프
                    import asyncio
                    wait_time = 2 ** attempt
                    logger.info(f"{wait_time}초 대기 후 재시도...")
                    await asyncio.sleep(wait_time)
                else:
                    logger.error(f"Playwright 게시판 조회 실패: {e}")
                    break
        
        # 게시판을 찾지 못한 경우 기본 게시판 생성
        if not boards:
            boards = self._generate_default_boards(cafe_info)
        
        return boards
    
    def _extract_board_id_from_url(self, url: str) -> str:
        """게시판 URL에서 게시판 ID 추출"""
        import re
        # 다양한 게시판 URL 패턴 매칭
        patterns = [
            r'menutype=(\d+)',
            r'boardtype=(\d+)', 
            r'menuid=(\d+)',
            r'/(\d+)$',
            r'clubid=\d+&boardtype=(\w+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        
        return "unknown"
    
    def _generate_default_boards(self, cafe_info: CafeInfo) -> List[BoardInfo]:
        """기본 게시판 목록 생성 (크롤링 실패 시)"""
        return [
            BoardInfo(
                name="전체게시글",
                url=f"{cafe_info.url}/ArticleList.nhn?search.clubid={cafe_info.cafe_id}&search.menuid=0",
                board_id="0",
                article_count=0
            ),
            BoardInfo(
                name="자유게시판",
                url=f"{cafe_info.url}/ArticleList.nhn?search.clubid={cafe_info.cafe_id}&search.menuid=1",
                board_id="1", 
                article_count=0
            )
        ]
    
    
    def export_users_to_excel(self, file_path: str, users: List[ExtractedUser]) -> bool:
        """사용자 목록을 엑셀로 내보내기 - CLAUDE.md: 파일 I/O는 adapters 담당"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 1. 엑셀 워크북 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "추출된 사용자"
            
            # 2. 헤더 작성 (형식화)
            headers = ["번호", "사용자 ID", "닉네임", "추출 시간"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 3. 데이터 정규화 및 작성
            for row, user in enumerate(users, 2):
                ws.cell(row=row, column=1, value=row-1)  # 번호
                ws.cell(row=row, column=2, value=user.user_id or "")  # 사용자 ID
                ws.cell(row=row, column=3, value=user.nickname or "")  # 닉네임
                # 날짜 정규화
                date_str = user.last_seen.strftime("%Y-%m-%d %H:%M:%S") if user.last_seen else ""
                ws.cell(row=row, column=4, value=date_str)  # 추출 시간
            
            # 4. 컬럼 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 5. 파일 저장 (실제 I/O)
            wb.save(file_path)
            logger.debug(f"엑셀 파일 저장 완료: {file_path} ({len(users)}개 레코드)")
            return True
            
        except Exception as e:
            logger.error(f"엑셀 파일 저장 실패: {e}")
            return False
    
    def export_users_to_meta_csv(self, file_path: str, users: List[ExtractedUser]) -> bool:
        """사용자 목록을 Meta CSV로 내보내기 - META_CSV_DOMAINS 기반 동적 생성"""
        try:
            # 1. 사용자 ID 정규화/중복제거
            all_user_ids = []
            for idx, user in enumerate(users):
                raw = (user.user_id or f"user{idx}")
                clean = "".join(c for c in raw if c.isalnum() or c == '_') or f"user{idx}"
                if clean not in all_user_ids:
                    all_user_ids.append(clean)
            all_user_ids.sort()

            if not all_user_ids:
                logger.warning("내보낼 사용자 ID가 없습니다")
                return False

            # 2. 헤더/행을 META_CSV_DOMAINS 기반으로 동적 생성
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                # 헤더
                writer.writerow(['email'] * len(META_CSV_DOMAINS))
                # 각 행: 도메인 개수만큼 이메일 컬럼
                for uid in all_user_ids:
                    row = [uid + domain for domain in META_CSV_DOMAINS]
                    writer.writerow(row)

            logger.debug(
                f"Meta CSV 파일 저장 완료: {file_path} "
                f"(사용자 {len(all_user_ids)}명 → 이메일 {len(all_user_ids)*len(META_CSV_DOMAINS)}개)"
            )
            return True

        except Exception as e:
            logger.error(f"Meta CSV 파일 저장 실패: {e}")
            return False
    
    async def fetch_sibling_articles(
        self,
        session,            # aiohttp.ClientSession 또는 동일 인터페이스
        clubid: str,
        articleid: str,
        boardtype: str,
        limit: int = 15,
        page: int = 1,
    ) -> Tuple[List[Dict], int]:
        """
        siblings API 호출 - CLAUDE.md: 벤더 API 호출은 adapters 담당
        Returns: (articles_items, api_calls)
        api_calls는 재시도/분할 호출 시 집계용으로 1 이상 반환 가능
        """
        await self.rate_limiter.wait_async()
        
        url = (
            f"https://apis.naver.com/cafe-web/cafe-articleapi/cafes/{clubid}/"
            f"articles/{articleid}/siblings?boardType={boardtype}&limit={limit}"
            f"&fromAllArticleList=false&filterByHeadId=false&page={page}&requestFrom=A"
        )
        try:
            async with session.get(url, timeout=15) as resp:
                if resp.status != 200:
                    logger.error(f"siblings API 응답코드: {resp.status} url={url}")
                    return [], 0
                data = await resp.json()
                items = data.get('articles', {}).get('items', [])
                logger.debug(f"siblings API 성공: {len(items)}개 아이템")
                return items, 1
        except Exception as e:
            logger.error(f"siblings API 호출 실패: {e}")
            return [], 0
    
    def cleanup_pages(self):
        """사용하던 페이지들 정리 - 비동기 안전"""
        try:
            import asyncio
            
            # 현재 이벤트 루프 확인
            try:
                loop = asyncio.get_running_loop()
                # 이미 실행 중인 루프가 있으면 태스크로 추가
                loop.create_task(self._async_cleanup_pages())
            except RuntimeError:
                # 실행 중인 루프가 없으면 새로 생성
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    loop.run_until_complete(self._async_cleanup_pages())
                finally:
                    loop.close()
                    
        except Exception as e:
            logger.warning(f"페이지 정리 중 오류: {e}")
    
    async def _async_cleanup_pages(self):
        """비동기 페이지 정리"""
        try:
            if hasattr(self, '_search_page') and self._search_page:
                await self._search_page.close()
                self._search_page = None
                logger.debug("검색 페이지 정리 완료")
        except Exception as e:
            logger.warning(f"검색 페이지 정리 오류: {e}")
            self._search_page = None
            
        try:
            if hasattr(self, '_board_page') and self._board_page:
                await self._board_page.close()
                self._board_page = None
                logger.debug("게시판 페이지 정리 완료")
        except Exception as e:
            logger.warning(f"게시판 페이지 정리 오류: {e}")
            self._board_page = None
            
        try:
            if hasattr(self, '_extraction_page') and self._extraction_page:
                await self._extraction_page.close()
                self._extraction_page = None
                logger.debug("추출 페이지 정리 완료")
        except Exception as e:
            logger.warning(f"추출 페이지 정리 오류: {e}")
            self._extraction_page = None
            
        try:
            if hasattr(self, '_info_page') and self._info_page:
                await self._info_page.close()
                self._info_page = None
                logger.debug("정보 페이지 정리 완료")
        except Exception as e:
            logger.warning(f"정보 페이지 정리 오류: {e}")
            self._info_page = None
            
        # aiohttp 세션 정리
        try:
            if hasattr(self, '_aiohttp_session') and self._aiohttp_session:
                await self._aiohttp_session.close()
                self._aiohttp_session = None
                logger.debug("aiohttp 세션 정리 완료")
        except Exception as e:
            logger.warning(f"aiohttp 세션 정리 오류: {e}")
            self._aiohttp_session = None


class RateLimiter:
    """Rate Limiting 헬퍼"""
    
    def __init__(self, requests_per_minute: int = 30):
        self.requests_per_minute = requests_per_minute
        self.min_interval = 60.0 / requests_per_minute
        self.last_request_time = 0.0
    
    def wait(self):
        """동기 환경에서 사용 (블로킹)"""
        current_time = time.time()
        time_since_last = current_time - self.last_request_time
        
        if time_since_last < self.min_interval:
            wait_time = self.min_interval - time_since_last
            logger.debug(f"Rate limiting: {wait_time:.2f}초 대기 (동기)")
            time.sleep(wait_time)
        
        self.last_request_time = time.time()
    
    async def wait_async(self):
        """비동기 환경에서 사용 (이벤트 루프 비블로킹)"""
        import asyncio
        current_time = time.time()
        time_since_last = current_time - self.last_request_time
        
        if time_since_last < self.min_interval:
            wait_time = self.min_interval - time_since_last
            logger.debug(f"Rate limiting: {wait_time:.2f}초 대기 (비동기)")
            await asyncio.sleep(wait_time)
        
        self.last_request_time = time.time()