"""
파워링크 분석기 어댁터
vendors의 Raw API 응답을 파워링크 분석 데이터로 변환 + 엑셀 내보내기
"""
import time
import random
import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle
from typing import List, Optional, Dict, Tuple
from playwright.sync_api import BrowserContext
from datetime import datetime

from src.vendors.naver.client_factory import get_keyword_tool_client
from src.foundation.logging import get_logger
from src.desktop.common_log import log_manager
from .models import KeywordAnalysisResult, BidPosition

# 파워링크 분석 설정
POWERLINK_CONFIG = {
    "pc_positions": 10,  # PC 입찰가 조회 위치 (1~10위)
    "mobile_positions": 5,  # 모바일 입찰가 조회 위치 (1~5위)
    "max_retries": 2,  # 크롤링 최대 재시도 횟수
    "request_timeout": 10,  # API 요청 타임아웃(초)
}

# 네이버 최소 입찰가
NAVER_MIN_BID = 70

logger = get_logger("features.powerlink_analyzer.adapters")


class AdaptiveRateLimiter:
    """적응형 Rate Limiter - API 응답에 따라 동적으로 호출 속도 조절"""
    
    def __init__(self, initial_delay=1.0, min_delay=0.5, max_delay=10.0):
        self.current_delay = initial_delay
        self.min_delay = min_delay
        self.max_delay = max_delay
        self.success_count = 0
        self.error_count = 0
        self.last_call_time = 0
        
    def on_success(self):
        """API 호출 성공 시 호출 - 점진적으로 속도 증가"""
        self.success_count += 1
        self.error_count = 0
        
        # 연속 성공 시 딜레이 감소 (최소값까지)
        if self.success_count >= 3:
            self.current_delay = max(self.min_delay, self.current_delay * 0.9)
            self.success_count = 0
            
    def on_rate_limit(self, retry_after=None):
        """429 Rate Limit 에러 시 호출"""
        self.error_count += 1
        self.success_count = 0
        
        if retry_after:
            # 서버에서 지정한 대기 시간 적용
            self.current_delay = min(float(retry_after), self.max_delay)
        else:
            # 딜레이 2배 증가
            self.current_delay = min(self.current_delay * 2, self.max_delay)
            
        logger.warning(f"Rate Limit 적응: 딜레이 {self.current_delay:.2f}초로 증가")
    
    def on_error(self):
        """일반적인 에러 시 호출"""
        self.error_count += 1
        if self.error_count >= 2:
            # 에러가 반복되면 딜레이 증가
            self.current_delay = min(self.current_delay * 1.5, self.max_delay)
            
    def wait(self):
        """API 호출 전 대기"""
        current_time = time.time()
        time_since_last = current_time - self.last_call_time
        
        if time_since_last < self.current_delay:
            wait_time = self.current_delay - time_since_last
            logger.debug(f"적응형 Rate Limit: {wait_time:.2f}초 대기")
            time.sleep(wait_time)
            
        self.last_call_time = time.time()


# 전역 적응형 Rate Limiter 인스턴스
adaptive_rate_limiter = AdaptiveRateLimiter()


def exponential_backoff_retry(func, max_retries=3, base_delay=1.0, max_delay=60.0, backoff_factor=2.0):
    """
    지수 백오프를 사용한 적응형 재시도
    
    Args:
        func: 실행할 함수
        max_retries: 최대 재시도 횟수
        base_delay: 기본 지연 시간 (초)
        max_delay: 최대 지연 시간 (초)
        backoff_factor: 백오프 배수
    """
    for attempt in range(max_retries):
        try:
            return func()
        except Exception as e:
            if attempt == max_retries - 1:  # 마지막 시도
                raise e
            
            # 지수 백오프 계산 (지터 추가)
            delay = min(base_delay * (backoff_factor ** attempt), max_delay)
            jitter = random.uniform(0, delay * 0.1)  # 10% 지터
            total_delay = delay + jitter
            
            logger.warning(f"재시도 {attempt + 1}/{max_retries} - {total_delay:.2f}초 후 재시도: {e}")
            time.sleep(total_delay)
    
    return None


class PowerLinkDataAdapter:
    """파워링크 분석을 위한 데이터 어댑터"""
    
    def __init__(self):
        """초기화"""
        self.playwright_helper = None
    
    def is_api_configured(self) -> bool:
        """API 설정 확인"""
        try:
            # foundation의 config_manager를 통해 설정 확인
            from src.foundation.config import config_manager
            api_config = config_manager.load_api_config()
            return api_config.is_searchad_valid()
        except Exception as e:
            logger.error(f"API 설정 확인 실패: {e}")
            return False
    
    def get_keyword_basic_data(self, keyword: str) -> Optional[Tuple[int, int, float, float, float, float]]:
        """
        기본 키워드 데이터 조회 (PC/모바일 검색량·클릭·CTR)

        Returns:
            (pc_search_volume, mobile_search_volume, pc_clicks, pc_ctr, mobile_clicks, mobile_ctr)
            또는 None
        """
        if not self.is_api_configured():
            logger.error("네이버 검색광고 API가 설정되지 않았습니다.")
            return None
        
        def _get_keyword_data():
            # 적응형 Rate Limiting 적용
            adaptive_rate_limiter.wait()
            
            try:
                # 키워드를 대문자로 변환하여 API 호출 (네이버 API 요구사항)
                api_keyword = keyword.upper()
                
                # vendors의 keyword_client 사용 (필요할 때마다 가져오기)
                keyword_client = get_keyword_tool_client()
                response_data = keyword_client.get_keyword_ideas([api_keyword], show_detail=True)
                
                # 성공 시 Rate Limiter 업데이트
                adaptive_rate_limiter.on_success()
                
                keyword_list = response_data.get('keywordList', [])
                for item in keyword_list:
                    # 대소문자 무관하게 비교
                    rel_keyword = item.get('relKeyword', '').upper()
                    if rel_keyword == api_keyword:
                        # PC & 모바일 검색량 처리
                        pc_search_volume = self._parse_search_volume(item.get('monthlyPcQcCnt'))
                        mobile_search_volume = self._parse_search_volume(item.get('monthlyMobileQcCnt'))
                        
                        # PC & 모바일 월평균 클릭수
                        pc_clicks = float(item.get('monthlyAvePcClkCnt', 0))
                        mobile_clicks = float(item.get('monthlyAveMobileClkCnt', 0))
                        
                        # 클릭률
                        pc_ctr = float(item.get('monthlyAvePcCtr', 0))
                        mobile_ctr = float(item.get('monthlyAveMobileCtr', 0))
                        
                        # PC/Mobile 검색량을 분리해서 반환 (더 정확한 순위 계산을 위해)
                        return (pc_search_volume, mobile_search_volume, pc_clicks, pc_ctr, mobile_clicks, mobile_ctr)
                
                # 키워드를 찾지 못한 경우
                return None
                
            except Exception as e:
                # API 에러 타입에 따른 Rate Limiter 업데이트
                error_msg = str(e).lower()
                if '429' in error_msg or 'rate limit' in error_msg:
                    # Retry-After 헤더 파싱 시도
                    retry_after = None
                    if hasattr(e, 'response') and e.response:
                        retry_after = e.response.headers.get('Retry-After')
                    adaptive_rate_limiter.on_rate_limit(retry_after)
                else:
                    adaptive_rate_limiter.on_error()
                raise
        
        # 적응형 재시도 적용
        try:
            return exponential_backoff_retry(_get_keyword_data, max_retries=3, base_delay=1.0)
        except Exception as e:
            logger.error(f"키워드 데이터 조회 최종 실패: {e}")
            return None
        
        return None
    
    def _parse_search_volume(self, volume_str: str) -> int:
        """검색량 문자열 파싱 ("< 10" 처리)"""
        if not volume_str or volume_str == "< 10":
            return 0
        try:
            return int(volume_str)
        except (ValueError, TypeError):
            return 0
    
    def get_bid_positions_for_both_devices(self, keyword: str) -> Tuple[List[BidPosition], List[BidPosition]]:
        """
        PC/모바일 입찰가를 각각 조회 (실제 API 테스트 확인: 통합 요청 불가, 따로 요청 필수)
        
        API 테스트 결과:
        - PC만 요청: 성공 ✅
        - 모바일만 요청: 성공 ✅  
        - device="ALL": 400 에러 ❌
        - device 없음: 400 에러 ❌
        
        Returns:
            (PC 입찰가 리스트, 모바일 입찰가 리스트)
        """
        if not self.is_api_configured():
            logger.error("네이버 검색광고 API가 설정되지 않았습니다.")
            return [], []
        
        # 병렬 처리를 위해 각각 조회 (순차 처리보다 빠름)
        pc_bids = self._get_bid_positions_for_device(keyword, "PC", POWERLINK_CONFIG["pc_positions"])
        mobile_bids = self._get_bid_positions_for_device(keyword, "MOBILE", POWERLINK_CONFIG["mobile_positions"])
        
        return pc_bids, mobile_bids
    
    def _get_bid_positions_for_device(self, keyword: str, device: str, max_positions: int) -> List[BidPosition]:
        """특정 디바이스의 입찰가 조회 - vendors 레이어를 통해 수행"""
        def _get_bid_data():
            # 적응형 Rate Limiting 적용
            adaptive_rate_limiter.wait()
            
            try:
                from src.vendors.naver.searchad.base_client import NaverKeywordToolClient
                
                # 베이스 클라이언트 인스턴스 생성 및 입찰가 조회
                client = NaverKeywordToolClient()
                positions_list = list(range(1, max_positions + 1))
                # 키워드를 대문자로 변환하여 API 호출 (네이버 API 요구사항)
                api_keyword = keyword.upper()
                response_data = client.get_bid_estimates(api_keyword, device, positions_list)
                
                # 성공 시 Rate Limiter 업데이트
                adaptive_rate_limiter.on_success()
                
                if not response_data:
                    raise Exception(f"{device} 입찰가 조회 실패: 응답 데이터 없음")
                    
                bid_positions = []
                estimates = response_data.get('estimate', [])
                for estimate in estimates:
                    bid_positions.append(BidPosition(
                        position=estimate['position'],
                        bid_price=estimate['bid']
                    ))
                return bid_positions
                
            except Exception as e:
                # API 에러 타입에 따른 Rate Limiter 업데이트
                error_msg = str(e).lower()
                if '429' in error_msg or 'rate limit' in error_msg:
                    retry_after = None
                    if hasattr(e, 'response') and e.response:
                        retry_after = e.response.headers.get('Retry-After')
                    adaptive_rate_limiter.on_rate_limit(retry_after)
                else:
                    adaptive_rate_limiter.on_error()
                raise
        
        # 적응형 재시도 적용
        try:
            return exponential_backoff_retry(_get_bid_data, max_retries=3, base_delay=1.0)
        except Exception as e:
            logger.error(f"{device} 입찰가 조회 최종 실패: {e}")
            return []
    
    
    def calculate_min_exposure_bid(self, bid_positions: List[BidPosition], position: int) -> int:
        """
        현실적인 최소노출가격 계산 (70원 API 오류 감지)
        """
        if not bid_positions or position <= 0:
            return 0
        if position > len(bid_positions):
            return bid_positions[-1].bid_price
        
        # 1페이지 노출 위치까지만 고려
        relevant_positions = bid_positions[:position]
        if not relevant_positions:
            return 0
        
        # 마지막 노출 위치 가격
        last_position_price = relevant_positions[-1].bid_price
        
        # 마지막 위치가 70원이고, 위에 정상 가격이 있는 경우 API 오류 감지
        if last_position_price == NAVER_MIN_BID and len(relevant_positions) > 1:
            # 위 순위들 중 70원이 아닌 가격들 확인
            non_min_prices = [pos.bid_price for pos in relevant_positions[:-1] 
                             if pos.bid_price > NAVER_MIN_BID]
            
            if non_min_prices:
                # 바로 위 순위 가격 확인
                second_last_price = relevant_positions[-2].bid_price
                
                # 바로 위 순위와 10배 이상 차이나면 API 오류로 판단
                if second_last_price / NAVER_MIN_BID > 10:
                    return second_last_price
                
                # 연속된 70원이 많으면 (절반 이상) 진짜 저경쟁 키워드
                min_bid_count = sum(1 for pos in relevant_positions 
                                  if pos.bid_price == NAVER_MIN_BID)
                
                if min_bid_count >= len(relevant_positions) // 2:
                    return NAVER_MIN_BID  # 진짜 저경쟁 키워드
                else:
                    return second_last_price  # API 오류 추정
        
        # 정상 경우는 해당 위치 가격 그대로 반환
        return last_position_price
    
    async def initialize_playwright_pages(self):
        """
        Playwright 브라우저 페이지 초기화 (vendors 호출 캡슐화)
        
        Returns:
            (playwright_helper, pc_page, mobile_page)
        """
        try:
            from src.vendors.web_automation.playwright_helper import create_playwright_helper, get_fast_browser_config
            
            # vendors의 최적화된 설정 활용
            config = get_fast_browser_config(headless=True)
            self.playwright_helper = await create_playwright_helper(config)
            
            # vendors 헬퍼의 context와 페이지 사용 (최적화 자동 적용됨)
            async_context = self.playwright_helper.context
            
            # PC/Mobile 페이지 생성 (vendors 최적화 자동 적용)
            pc_page = await async_context.new_page()
            mobile_page = await async_context.new_page()
            
            logger.info("PC/Mobile 페이지 초기화 완료 (vendors 헬퍼 활용)")
            return self.playwright_helper, pc_page, mobile_page
            
        except Exception as e:
            logger.error(f"페이지 초기화 실패: {e}")
            raise
    
    async def cleanup_playwright(self, playwright_helper, pc_page=None, mobile_page=None):
        """
        Playwright 리소스 정리 (vendors 호출 캡슐화)
        """
        try:
            # 개별 페이지 정리
            if pc_page:
                await pc_page.close()
            if mobile_page:
                await mobile_page.close()
            
            # vendors 헬퍼 정리 (context, browser, playwright 모두 정리됨)
            if playwright_helper:
                await playwright_helper.cleanup()
                
            logger.info("페이지 정리 완료")
        except Exception as e:
            logger.warning(f"페이지 정리 중 오류: {e}")


class PowerLinkExcelExporter:
    """PowerLink 분석 결과 엑셀 내보내기 클래스"""
    
    def __init__(self):
        pass
    
    def _setup_number_formats(self, workbook):
        """숫자 포맷 스타일 설정"""
        # 천 단위 콤마 스타일 (숫자로 저장하되 콤마 표시)
        if 'number_comma' not in workbook.named_styles:
            number_style = NamedStyle(name='number_comma')
            number_style.number_format = '#,##0'
            workbook.add_named_style(number_style)
        
        # 퍼센트 스타일
        if 'percent_style' not in workbook.named_styles:
            percent_style = NamedStyle(name='percent_style')
            percent_style.number_format = '0.00"%"'
            workbook.add_named_style(percent_style)
        
        return workbook
    
    def export_to_excel(self, keywords_data, file_path: str, session_name: str = ""):
        """
        PowerLink 분석 결과를 엑셀 파일로 내보내기
        
        Args:
            keywords_data: 키워드 분석 결과 딕셔너리 (KeywordAnalysisResult 또는 dict)
            file_path: 저장할 엑셀 파일 경로
            session_name: 세션명 (로그용)
        """
        try:
            if not keywords_data:
                raise Exception("저장할 분석 데이터가 없습니다.")
            
            # 데이터 정규화 (dict인 경우 KeywordAnalysisResult로 변환)
            normalized_data = {}
            for keyword, data in keywords_data.items():
                if isinstance(data, dict):
                    # dict인 경우 KeywordAnalysisResult 객체로 변환
                    normalized_data[keyword] = self._dict_to_result(data)
                else:
                    # 이미 KeywordAnalysisResult인 경우 그대로 사용
                    normalized_data[keyword] = data
            
            workbook = openpyxl.Workbook()
            
            # 숫자 포맷 스타일 설정
            workbook = self._setup_number_formats(workbook)
            
            # 모바일 분석결과 시트 생성
            self._create_mobile_sheet(workbook, normalized_data)
            
            # PC 분석결과 시트 생성
            self._create_pc_sheet(workbook, normalized_data)
            
            # 파일 저장
            workbook.save(file_path)
            
            log_message = f"PowerLink 엑셀 파일 생성 완료"
            if session_name:
                log_message += f" ({session_name})"
            log_message += f": {file_path}"
            
            log_manager.add_log(log_message, "success")
            logger.info(log_message)
            
        except Exception as e:
            error_message = f"PowerLink 엑셀 파일 생성 실패: {str(e)}"
            log_manager.add_log(error_message, "error")
            logger.error(error_message)
            raise
    
    def _create_mobile_sheet(self, workbook: openpyxl.Workbook, keywords_data: Dict[str, KeywordAnalysisResult]):
        """모바일 분석결과 시트 생성"""
        # 첫 번째 시트를 모바일로 설정
        mobile_sheet = workbook.active
        mobile_sheet.title = "모바일분석결과"
        
        # 모바일 헤더 설정 (5위까지 순위별 광고비 포함)
        mobile_headers = [
            "키워드", "월검색량", "모바일클릭수", "모바일클릭률", 
            "1p노출수", "1등광고비", "최소노출가격", "추천순위", ""  # 빈 칸
        ]
        
        # 1위부터 5위까지 광고비 헤더 추가
        for i in range(1, 6):
            mobile_headers.append(f"{i}위광고비")
        
        # 헤더 스타일 적용
        for col, header in enumerate(mobile_headers, 1):
            cell = mobile_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 모바일 데이터 추가 (추천순위 순으로 정렬)
        sorted_mobile = sorted(
            keywords_data.values(), 
            key=lambda x: x.mobile_recommendation_rank if x.mobile_recommendation_rank > 0 else 999
        )
        
        for row_idx, result in enumerate(sorted_mobile, 2):
            # 기본 데이터
            mobile_sheet.cell(row=row_idx, column=1, value=result.keyword)
            
            # 월검색량 (Mobile 검색량, 숫자로 저장, 콤마 표시)
            monthly_volume_cell = mobile_sheet.cell(row=row_idx, column=2, value=result.mobile_search_volume)
            monthly_volume_cell.style = 'number_comma'
            
            mobile_sheet.cell(row=row_idx, column=3, value=result.mobile_clicks)
            
            # 클릭률 (숫자로 저장, 퍼센트 표시)
            ctr_cell = mobile_sheet.cell(row=row_idx, column=4, value=result.mobile_ctr)
            ctr_cell.style = 'percent_style'
            
            mobile_sheet.cell(row=row_idx, column=5, value=result.mobile_first_page_positions)
            
            # 1등광고비 (숫자로 저장, 콤마 표시)
            first_bid_cell = mobile_sheet.cell(row=row_idx, column=6, value=result.mobile_first_position_bid)
            first_bid_cell.style = 'number_comma'
            
            # 최소노출가격 (숫자로 저장, 콤마 표시)
            min_bid_cell = mobile_sheet.cell(row=row_idx, column=7, value=result.mobile_min_exposure_bid)
            min_bid_cell.style = 'number_comma'
            
            # 추천순위 (숫자로 저장)
            if result.mobile_recommendation_rank > 0:
                mobile_sheet.cell(row=row_idx, column=8, value=result.mobile_recommendation_rank)
            else:
                mobile_sheet.cell(row=row_idx, column=8, value="-")
            
            # 9번 컬럼은 빈 칸 (건너뜀)
            
            # 10번 컬럼부터 순위별 입찰가 (5위까지만) - 숫자로 저장, 콤마 표시
            if hasattr(result, 'mobile_bid_positions') and result.mobile_bid_positions:
                for idx, bid_pos in enumerate(result.mobile_bid_positions[:5]):
                    bid_cell = mobile_sheet.cell(row=row_idx, column=10+idx, value=bid_pos.bid_price)
                    bid_cell.style = 'number_comma'
        
        # 컬럼 너비 자동 조정
        self._auto_adjust_columns(mobile_sheet)
    
    def _create_pc_sheet(self, workbook: openpyxl.Workbook, keywords_data: Dict[str, KeywordAnalysisResult]):
        """PC 분석결과 시트 생성"""
        # PC 시트 생성
        pc_sheet = workbook.create_sheet("PC분석결과")
        
        # PC 헤더 설정 (10위까지 순위별 광고비 포함)
        pc_headers = [
            "키워드", "월검색량", "PC클릭수", "PC클릭률", 
            "1p노출수", "1등광고비", "최소노출가격", "추천순위", ""  # 빈 칸
        ]
        
        # 1위부터 10위까지 광고비 헤더 추가
        for i in range(1, 11):
            pc_headers.append(f"{i}위광고비")
        
        # 헤더 스타일 적용
        for col, header in enumerate(pc_headers, 1):
            cell = pc_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # PC 데이터 추가 (추천순위 순으로 정렬)
        sorted_pc = sorted(
            keywords_data.values(), 
            key=lambda x: x.pc_recommendation_rank if x.pc_recommendation_rank > 0 else 999
        )
        
        for row_idx, result in enumerate(sorted_pc, 2):
            # 기본 데이터
            pc_sheet.cell(row=row_idx, column=1, value=result.keyword)
            
            # 월검색량 (PC 검색량, 숫자로 저장, 콤마 표시)
            monthly_volume_cell = pc_sheet.cell(row=row_idx, column=2, value=result.pc_search_volume)
            monthly_volume_cell.style = 'number_comma'
            
            pc_sheet.cell(row=row_idx, column=3, value=result.pc_clicks)
            
            # 클릭률 (숫자로 저장, 퍼센트 표시)
            ctr_cell = pc_sheet.cell(row=row_idx, column=4, value=result.pc_ctr)
            ctr_cell.style = 'percent_style'
            
            pc_sheet.cell(row=row_idx, column=5, value=result.pc_first_page_positions)
            
            # 1등광고비 (숫자로 저장, 콤마 표시)
            first_bid_cell = pc_sheet.cell(row=row_idx, column=6, value=result.pc_first_position_bid)
            first_bid_cell.style = 'number_comma'
            
            # 최소노출가격 (숫자로 저장, 콤마 표시)
            min_bid_cell = pc_sheet.cell(row=row_idx, column=7, value=result.pc_min_exposure_bid)
            min_bid_cell.style = 'number_comma'
            
            # 추천순위 (숫자로 저장)
            if result.pc_recommendation_rank > 0:
                pc_sheet.cell(row=row_idx, column=8, value=result.pc_recommendation_rank)
            else:
                pc_sheet.cell(row=row_idx, column=8, value="-")
            
            # 9번 컬럼은 빈 칸 (건너뜀)
            
            # 10번 컬럼부터 순위별 입찰가 (10위까지만) - 숫자로 저장, 콤마 표시
            if hasattr(result, 'pc_bid_positions') and result.pc_bid_positions:
                for idx, bid_pos in enumerate(result.pc_bid_positions[:10]):
                    bid_cell = pc_sheet.cell(row=row_idx, column=10+idx, value=bid_pos.bid_price)
                    bid_cell.style = 'number_comma'
        
        # 컬럼 너비 자동 조정
        self._auto_adjust_columns(pc_sheet)
    
    def _auto_adjust_columns(self, sheet):
        """컬럼 너비 자동 조정"""
        try:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 최소 10, 최대 30 범위로 제한
                adjusted_width = min(max(max_length + 2, 10), 30)
                sheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            logger.warning(f"컬럼 너비 자동 조정 실패: {e}")
    
    def _dict_to_result(self, data_dict: dict) -> KeywordAnalysisResult:
        """딕셔너리를 KeywordAnalysisResult 객체로 변환"""
        # BidPosition 객체들 복원
        def restore_bid_positions(bid_data):
            if not bid_data:
                return []
            
            positions = []
            for item in bid_data:
                if isinstance(item, dict):
                    positions.append(BidPosition(
                        position=item.get('position', 0),
                        bid_price=item.get('bid_price', 0)
                    ))
                else:
                    # 이미 BidPosition 객체인 경우
                    positions.append(item)
            return positions
        
        # analyzed_at 필드 처리
        analyzed_at = data_dict.get('analyzed_at')
        if analyzed_at and isinstance(analyzed_at, str):
            try:
                analyzed_at = datetime.fromisoformat(analyzed_at)
            except:
                analyzed_at = datetime.now()
        elif not analyzed_at:
            analyzed_at = datetime.now()
        
        return KeywordAnalysisResult(
            keyword=data_dict.get('keyword', ''),
            pc_search_volume=data_dict.get('pc_search_volume', 0),
            mobile_search_volume=data_dict.get('mobile_search_volume', 0),
            pc_clicks=data_dict.get('pc_clicks', 0.0),
            pc_ctr=data_dict.get('pc_ctr', 0.0),
            pc_first_page_positions=data_dict.get('pc_first_page_positions', 0),
            pc_first_position_bid=data_dict.get('pc_first_position_bid', 0),
            pc_min_exposure_bid=data_dict.get('pc_min_exposure_bid', 0),
            pc_bid_positions=restore_bid_positions(data_dict.get('pc_bid_positions', [])),
            mobile_clicks=data_dict.get('mobile_clicks', 0.0),
            mobile_ctr=data_dict.get('mobile_ctr', 0.0),
            mobile_first_page_positions=data_dict.get('mobile_first_page_positions', 0),
            mobile_first_position_bid=data_dict.get('mobile_first_position_bid', 0),
            mobile_min_exposure_bid=data_dict.get('mobile_min_exposure_bid', 0),
            mobile_bid_positions=restore_bid_positions(data_dict.get('mobile_bid_positions', [])),
            pc_recommendation_rank=data_dict.get('pc_recommendation_rank', 0),
            mobile_recommendation_rank=data_dict.get('mobile_recommendation_rank', 0),
            analyzed_at=analyzed_at
        )


class HistoryExportAdapter:
    """히스토리 내보내기 전용 어댑터 (파일 I/O 담당)"""
    
    def __init__(self):
        pass
    
    def choose_single_file_path(self, session_data: dict, parent_widget=None) -> tuple[bool, str]:
        """
        단일 세션 저장 경로 선택 (UI I/O 헬퍼)
        
        Args:
            session_data: {'id': int, 'name': str, 'created_at': str}
            parent_widget: 부모 위젯 (다이얼로그 표시용)
            
        Returns:
            (성공 여부, 선택된 파일 경로 또는 메시지)
        """
        try:
            from datetime import datetime
            from PySide6.QtWidgets import QFileDialog
            
            # 기본 파일명 생성 (세션 시간 기반)
            session_time = datetime.fromisoformat(session_data['created_at'])
            time_str = session_time.strftime('%Y%m%d_%H%M%S')
            default_filename = f"파워링크광고비분석_{time_str}.xlsx"
            
            # 파일 저장 다이얼로그
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "엑셀 파일 저장",
                default_filename,
                "Excel files (*.xlsx);;All files (*.*)"
            )
            
            return (bool(file_path), file_path or "사용자가 취소했습니다.")
                
        except Exception as e:
            logger.error(f"파일 경로 선택 실패: {e}")
            return False, str(e)
    
    def choose_output_folder(self, parent_widget=None) -> tuple[bool, str]:
        """
        출력 폴더 선택 (UI I/O 헬퍼)
        
        Args:
            parent_widget: 부모 위젯 (다이얼로그 표시용)
            
        Returns:
            (성공 여부, 선택된 폴더 경로)
        """
        try:
            from PySide6.QtWidgets import QFileDialog
            
            # 폴더 선택 다이얼로그
            folder_path = QFileDialog.getExistingDirectory(
                parent_widget,
                "엑셀 파일 저장 폴더 선택",
                ""
            )
            
            return (bool(folder_path), folder_path or "")
            
        except Exception as e:
            logger.error(f"폴더 선택 실패: {e}")
            return False, ""
    
    def show_export_success_dialog(self, file_path: str, file_count: int = 1, parent_widget=None, reference_widget=None):
        """엑셀 내보내기 성공 다이얼로그 표시"""
        try:
            from src.toolbox.ui_kit.modern_dialog import ModernSaveCompletionDialog
            
            if file_count == 1:
                title = "저장 완료"
                message = "엑셀 파일이 성공적으로 저장되었습니다."
            else:
                title = "선택저장 완료"
                message = f"{file_count}개의 엑셀 파일이 성공적으로 저장되었습니다."
            
            success_dialog = ModernSaveCompletionDialog(
                parent=parent_widget,
                title=title,
                message=message,
                file_path=file_path
            )
            
            # 위치 지정 제거 - 중앙에 표시되도록 함
            success_dialog.exec()
            
        except Exception as e:
            logger.error(f"성공 다이얼로그 표시 실패: {e}")
    
    def show_export_error_dialog(self, error_message: str, parent_widget=None):
        """엑셀 내보내기 실패 다이얼로그 표시"""
        try:
            from src.toolbox.ui_kit.modern_dialog import ModernConfirmDialog
            
            dialog = ModernConfirmDialog(
                parent_widget,
                "저장 실패",
                f"엑셀 파일 저장 중 오류가 발생했습니다.\n\n{error_message}",
                confirm_text="확인",
                cancel_text=None,
                icon="❌"
            )
            dialog.exec()
            
        except Exception as e:
            logger.error(f"에러 다이얼로그 표시 실패: {e}")


class CurrentAnalysisExportAdapter:
    """현재 분석 결과 내보내기 전용 어댑터 (파일 I/O 담당)"""
    
    def __init__(self):
        pass
    
    def export_current_analysis_with_dialog(self, keywords_data: dict, session_name: str = "", parent_widget=None) -> bool:
        """
        현재 분석 결과를 엑셀로 내보내기 (파일 다이얼로그 포함)
        
        Args:
            keywords_data: 키워드 분석 결과 딕셔너리
            session_name: 세션명
            parent_widget: 부모 위젯 (다이얼로그 표시용)
            
        Returns:
            성공 여부
        """
        try:
            from datetime import datetime
            from PySide6.QtWidgets import QFileDialog
            
            # 기본 파일명 생성
            time_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"파워링크광고비분석_{time_str}.xlsx"
            
            # 파일 저장 다이얼로그
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "엑셀 파일 저장",
                default_filename,
                "Excel files (*.xlsx)"
            )
            
            if not file_path:
                return False  # 사용자가 취소
            
            # 엑셀 파일 생성
            powerlink_excel_exporter.export_to_excel(keywords_data, file_path, session_name)
            
            # 성공 다이얼로그 표시
            self.show_current_export_success_dialog(file_path, parent_widget)
            
            return True
            
        except Exception as e:
            logger.error(f"현재 분석 결과 내보내기 실패: {e}")
            self.show_current_export_error_dialog(str(e), parent_widget)
            return False
    
    def show_current_export_success_dialog(self, file_path: str, parent_widget=None):
        """현재 분석 결과 엑셀 내보내기 성공 다이얼로그 표시"""
        try:
            from src.toolbox.ui_kit.modern_dialog import ModernSaveCompletionDialog
            from pathlib import Path
            
            filename = Path(file_path).name
            ModernSaveCompletionDialog.show_save_completion(
                parent=parent_widget,
                title="저장 완료",
                message=f"엑셀 파일이 성공적으로 저장되었습니다.\n\n파일명: {filename}",
                file_path=file_path
            )
            
        except Exception as e:
            logger.error(f"성공 다이얼로그 표시 실패: {e}")
    
    def show_current_export_error_dialog(self, error_message: str, parent_widget=None):
        """현재 분석 결과 엑셀 내보내기 실패 다이얼로그 표시"""
        try:
            from src.toolbox.ui_kit.modern_dialog import ModernConfirmDialog
            
            dialog = ModernConfirmDialog(
                parent_widget,
                "저장 실패",
                f"엑셀 파일 저장 중 오류가 발생했습니다.\n\n{error_message}",
                confirm_text="확인",
                cancel_text=None,
                icon="❌"
            )
            dialog.exec()
            
        except Exception as e:
            logger.error(f"에러 다이얼로그 표시 실패: {e}")


# 전역 익스포터 인스턴스
powerlink_excel_exporter = PowerLinkExcelExporter()
history_export_adapter = HistoryExportAdapter()
current_analysis_export_adapter = CurrentAnalysisExportAdapter()


